#!/usr/bin/env python3
"""Harvest public TCIA NIfTI companion metadata into a local SQLite database.

The first pass intentionally keeps heterogeneous spreadsheet rows intact while
also extracting a best-effort IDC-like series metadata shape where columns match
common names used in TCIA submissions.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import io
import json
import mimetypes
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any, Iterable, Optional


SKILL_ROOT = Path(__file__).resolve().parents[1]
SCHEMA_VERSION = 2
DEFAULT_SOURCE_DB = SKILL_ROOT / "cache" / "tcia_snapshot.sqlite"
DEFAULT_OUT_DIR = SKILL_ROOT / "outputs" / "nifti_metadata"
DEFAULT_OUT_DB = DEFAULT_OUT_DIR / "nifti_metadata.sqlite"
USER_AGENT = "tcia-nifti-metadata-harvest/0.1"
SPREADSHEET_EXTENSIONS = {".csv", ".tsv", ".xlsx"}
ZIP_EXTENSIONS = {".zip"}
NIFTI_EXTENSIONS = {".nii", ".nii.gz"}
NON_FILE_NIFTI_VALUES = {"directory", "dir", "folder", "file", "symbolic_link", "symlink"}
LOCAL_EXTRACTED_PACKAGE_SPECS = (
    {
        "short_title": "BraTS-TCGA-GBM",
        "dirname": "Pre-operative_TCGA_GBM_NIfTI_and_Segmentations",
    },
    {
        "short_title": "BraTS-TCGA-LGG",
        "dirname": "Pre-operative_TCGA_LGG_NIfTI_and_Segmentations",
    },
)
PANCREAS_CT_MANIFEST_FILE = "Pancreas-CT-20200910.tcia"
METADATA_NAME_HINTS = (
    "acquisition",
    "clinical",
    "data_dictionary",
    "datadictionary",
    "dictionary",
    "file",
    "image",
    "imaging",
    "mapping",
    "metadata",
    "radiology",
    "scanner",
    "sequence",
    "series",
    "spreadsheet",
)
SKIP_NAME_HINTS = (
    "radiomic",
    "feature",
    "pyradiomics",
)


FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "PatientID": (
        "patientid",
        "patient_id",
        "pid",
        "subjectid",
        "subject_id",
        "caseid",
        "case_id",
        "tcia_case_id",
        "tcia_caseid",
        "tcia_patient_id",
        "patientidontciaradiologyportal",
        "bratssubjectid",
        "tciaid",
        "tcia_id",
        "bratsid",
        "brats_id",
    ),
    "PatientAge": ("patientage", "patient_age", "age", "ageatimaging", "age_at_imaging"),
    "PatientSex": ("patientsex", "patient_sex", "sex", "gender"),
    "StudyInstanceUID": (
        "studyinstanceuid",
        "study_instance_uid",
        "studyuid",
        "tcia_study_instance_uid",
        "tciastudyinstanceuid",
    ),
    "StudyDate": ("studydate", "study_date", "scan_date", "exam_date", "date"),
    "StudyDescription": ("studydescription", "study_description", "exam_description"),
    "SeriesInstanceUID": (
        "seriesinstanceuid",
        "series_instance_uid",
        "seriesuid",
        "tcia_series_instance_uid",
        "tciaseriesinstanceuid",
    ),
    "SeriesDate": ("seriesdate", "series_date"),
    "SeriesDescription": (
        "seriesdescription",
        "series_description",
        "sequencedescription",
        "sequence_description",
        "sequence",
        "image_type",
        "imagetype",
    ),
    "SeriesNumber": ("seriesnumber", "series_number"),
    "BodyPartExamined": ("bodypartexamined", "body_part_examined", "anatomy", "bodypart"),
    "Modality": ("modality", "image_modality", "scan_modality"),
    "Manufacturer": ("manufacturer", "scanner_manufacturer", "vendor"),
    "ManufacturerModelName": (
        "manufacturermodelname",
        "manufacturer_model_name",
        "scanner",
        "scanner_model",
        "model",
    ),
    "SOPClassUID": ("sopclassuid", "sop_class_uid"),
    "sop_class_name": ("sopclassname", "sop_class_name"),
    "TransferSyntaxUID": ("transfersyntaxuid", "transfer_syntax_uid"),
    "transfer_syntax_name": ("transfersyntaxname", "transfer_syntax_name"),
    "instanceCount": ("instancecount", "instance_count", "images", "image_count", "number_of_images"),
    "series_size_MB": ("series_size_mb", "size_mb", "filesize_mb", "file_size_mb"),
    "nifti_file": (
        "nifti_file",
        "nifti_filename",
        "file",
        "filename",
        "file_name",
        "image_file",
        "image_filename",
        "path",
        "pathname",
    ),
    "MagneticFieldStrength": (
        "magneticfieldstrength",
        "magnetic_field_strength",
        "fieldstrength",
        "field_strength",
        "tesla",
    ),
    "ScanningSequence": ("scanningsequence", "scanning_sequence"),
    "SequenceVariant": ("sequencevariant", "sequence_variant"),
    "MRAcquisitionType": ("mracquisitiontype", "mr_acquisition_type", "acquisition_type"),
    "EchoTime": ("echotime", "echo_time", "echotimems", "echo_time_ms", "te"),
    "RepetitionTime": (
        "repetitiontime",
        "repetition_time",
        "repetitiontimems",
        "repetition_time_ms",
        "tr",
    ),
    "EchoTrainLength": ("echotrainlength", "echo_train_length"),
    "FlipAngle": ("flipangle", "flip_angle"),
    "InversionTime": (
        "inversiontime",
        "inversion_time",
        "inversiontimems",
        "inversion_time_ms",
        "ti",
    ),
    "ReceiveCoilName": ("receivecoilname", "receive_coil_name", "coil", "coil_name"),
    "SequenceName": ("sequencename", "sequence_name"),
    "DiffusionBValue": ("diffusionbvalue", "diffusion_b_value", "bvalue", "b_value"),
    "NumberOfTemporalPositions": (
        "numberoftemporalpositions",
        "number_of_temporal_positions",
        "temporal_positions",
    ),
    "PixelSpacing_row_mm": ("pixelspacingrowmm", "pixel_spacing_row_mm", "pixelspacingrow"),
    "PixelSpacing_col_mm": ("pixelspacingcolmm", "pixel_spacing_col_mm", "pixelspacingcol"),
    "Rows": ("rows", "height", "matrix_rows"),
    "Columns": ("columns", "cols", "width", "matrix_columns"),
    "SliceThickness": ("slicethickness", "slice_thickness", "slicethicknessmm", "slice_thickness_mm"),
    "KVP": ("kvp", "kv"),
    "ScanOptions": ("scanoptions", "scan_options"),
    "ConvolutionKernel": ("convolutionkernel", "convolution_kernel", "kernel"),
    "GantryDetectorTilt": ("gantrydetectortilt", "gantry_detector_tilt"),
    "XRayTubeCurrent_min": ("xraytubecurrentmin", "xray_tube_current_min"),
    "XRayTubeCurrent_max": ("xraytubecurrentmax", "xray_tube_current_max"),
    "Exposure_min": ("exposuremin", "exposure_min"),
    "Exposure_max": ("exposuremax", "exposure_max"),
    "ExposureTime_min": ("exposuretimemin", "exposure_time_min"),
    "ExposureTime_max": ("exposuretimemax", "exposure_time_max"),
    "DataCollectionDiameter": ("datacollectiondiameter", "data_collection_diameter"),
    "ReconstructionDiameter": ("reconstructiondiameter", "reconstruction_diameter"),
    "SpiralPitchFactor": ("spiralpitchfactor", "spiral_pitch_factor", "pitch"),
    "ContrastBolusAgent": ("contrastbolusagent", "contrast_bolus_agent", "contrast_agent"),
    "ContrastBolusIngredient": (
        "contrastbolusingredient",
        "contrast_bolus_ingredient",
        "contrast_ingredient",
    ),
    "ContrastBolusRoute": ("contrastbolusroute", "contrast_bolus_route", "contrast_route"),
    "SegmentationType": ("segmentationtype", "segmentation_type"),
    "total_segments": ("totalsegments", "total_segments", "segment_count"),
    "AlgorithmType": ("algorithmtype", "algorithm_type"),
    "AlgorithmName": ("algorithmname", "algorithm_name"),
    "segmented_SeriesInstanceUID": (
        "segmentedseriesinstanceuid",
        "segmented_series_instance_uid",
        "referencedseriesinstanceuid",
        "referenced_series_instance_uid",
        "tcia_series_instance_uid",
        "tciaseriesinstanceuid",
    ),
}


NORMALIZED_COLUMNS = [
    "collection_id",
    "analysis_result_id",
    "PatientID",
    "SeriesInstanceUID",
    "StudyInstanceUID",
    "source_DOI",
    "PatientAge",
    "PatientSex",
    "StudyDate",
    "StudyDescription",
    "BodyPartExamined",
    "Modality",
    "SOPClassUID",
    "sop_class_name",
    "TransferSyntaxUID",
    "transfer_syntax_name",
    "Manufacturer",
    "ManufacturerModelName",
    "SeriesDate",
    "SeriesDescription",
    "SeriesNumber",
    "instanceCount",
    "license_short_name",
    "series_init_idc_version",
    "series_revised_idc_version",
    "aws_bucket",
    "crdc_series_uuid",
    "series_aws_url",
    "series_size_MB",
    "nifti_file",
    "MagneticFieldStrength",
    "ScanningSequence",
    "SequenceVariant",
    "MRAcquisitionType",
    "EchoTime",
    "RepetitionTime",
    "EchoTrainLength",
    "FlipAngle",
    "InversionTime",
    "ReceiveCoilName",
    "SequenceName",
    "DiffusionBValue",
    "NumberOfTemporalPositions",
    "PixelSpacing_row_mm",
    "PixelSpacing_col_mm",
    "Rows",
    "Columns",
    "SliceThickness",
    "KVP",
    "ScanOptions",
    "ConvolutionKernel",
    "GantryDetectorTilt",
    "XRayTubeCurrent_min",
    "XRayTubeCurrent_max",
    "FilterType",
    "Exposure_min",
    "Exposure_max",
    "ExposureTime_min",
    "ExposureTime_max",
    "DataCollectionDiameter",
    "ReconstructionDiameter",
    "SpiralPitchFactor",
    "ContrastBolusAgent",
    "ContrastBolusIngredient",
    "ContrastBolusRoute",
    "SegmentationType",
    "total_segments",
    "AlgorithmType",
    "AlgorithmName",
    "segmented_SeriesInstanceUID",
]


def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def safe_slug(value: str, fallback: str = "item") -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip()).strip("._-")
    return slug or fallback


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def extension_for_name(name: str) -> str:
    lower = name.lower()
    if lower.endswith(".nii.gz"):
        return ".nii.gz"
    return Path(urllib.parse.urlparse(name).path).suffix.lower()


def clean_package_path(value: str) -> str:
    return str(value or "").strip().lstrip("/")


def split_package_path(value: str) -> list[str]:
    return [
        part.strip()
        for part in clean_package_path(value).split("/")
        if part and part.strip()
    ]


def strip_session_suffix(value: str) -> str:
    text = value.strip()
    for pattern in (
        r"^(.+?)[_-](?:timepoint|tp|t)\d+$",
        r"^(.+?)[_-]\d{1,2}$",
        r"^(.+?)[_-]day\d+$",
    ):
        match = re.fullmatch(pattern, text, flags=re.IGNORECASE)
        if match and re.search(r"\d{3,}", match.group(1)):
            return match.group(1)
    return text


def path_component_looks_like_subject_id(value: str) -> bool:
    text = value.strip("._- ")
    if not text:
        return False
    lower = text.lower()
    skip_exact = {
        "images",
        "masks",
        "mask",
        "nifti",
        "niftis",
        "seg",
        "segs",
        "segmentations",
        "segmentation",
        "training",
        "validation",
        "test",
        "labels",
        "resampled",
        "annotations",
        "atlas",
        "otherneoplasms",
    }
    if lower in skip_exact or lower.endswith("_nifti") or lower.endswith("-nifti"):
        return False
    if re.search(r"(?:^|[_-])(?:t|tp|timepoint|day)\d+$", lower):
        return False
    patterns = (
        r"[A-Za-z][A-Za-z0-9]*(?:[-_][A-Za-z0-9]+)+[-_]\d{1,6}(?:[-_]\d{1,6})?",
        r"(?:PatientID|Myel|id|BT|VS|YG)[-_]?[A-Za-z0-9]{3,}",
        r"\d{3,}",
    )
    return any(re.fullmatch(pattern, text) for pattern in patterns)


def path_component_looks_like_package_root(value: str, short_title: str) -> bool:
    text = value.strip("._- ")
    if not text:
        return False
    normalized_text = normalize_key(re.sub(r"(?:[_-]?v(?:ersion)?[_-]?\d+)$", "", text, flags=re.IGNORECASE))
    normalized_short_title = normalize_key(short_title)
    return bool(normalized_short_title and normalized_text.startswith(normalized_short_title))


def subject_id_from_file_stem(value: str) -> str:
    stem = value.strip("._- ")
    if not stem:
        return ""

    stem = re.sub(r"[_-]day\d+.*", "", stem, flags=re.IGNORECASE)
    patterns = (
        r"^(ACRIN[-_]\d+[-_]\d+)",
        r"^(BraTS[-_][A-Za-z0-9]+[-_]\d{3,}(?:[-_]\d{3,})?)",
        r"^(ISPY1[-_]\d+)",
        r"^(RHUH[-_]\d{4})",
        r"^(volume[-_]covid19[-_][A-Za-z][-_]\d+)",
    )
    for pattern in patterns:
        match = re.match(pattern, stem, flags=re.IGNORECASE)
        if match:
            return strip_session_suffix(match.group(1))

    if re.search(r"(?:adc|asl|brain|ce|dce|defaced|flair|label|mask|post|seg|t1|t1c|t2)", stem, flags=re.IGNORECASE):
        return ""
    if path_component_looks_like_subject_id(stem):
        return strip_session_suffix(stem)
    return ""


def infer_patient_id_from_path(short_title: str, package_path: str) -> tuple[str, str]:
    """Infer a patient-like identifier from common NIfTI package path layouts.

    This is intentionally conservative and only fills gaps when no submitter
    spreadsheet supplied a PatientID. The method string is stored in provenance.
    """

    parts = split_package_path(package_path)
    if not parts:
        return "", ""

    for index, part in enumerate(parts[:-1]):
        if index == 0 and path_component_looks_like_package_root(part, short_title):
            continue
        candidate = part.removesuffix("_nifti").removesuffix("-nifti")
        if path_component_looks_like_subject_id(candidate):
            return strip_session_suffix(candidate), "path_component"

    stem = strip_known_file_suffix(parts[-1])
    patient_id = subject_id_from_file_stem(stem)
    if patient_id:
        return patient_id, "file_stem"

    short_tokens = [token for token in re.split(r"[^A-Za-z0-9]+", short_title) if token]
    if short_tokens:
        prefix_pattern = r"[-_ ]+".join(re.escape(token) for token in short_tokens[:3])
        match = re.search(
            rf"\b({prefix_pattern}[-_ ]+\d{{1,6}}(?:[-_]\d{{1,6}})?)\b",
            package_path,
            flags=re.IGNORECASE,
        )
        if match:
            return strip_session_suffix(match.group(1).replace(" ", "-")), "short_title_pattern"

    return "", ""


def is_plausible_nifti_file(value: str) -> bool:
    text = clean_package_path(value)
    lower = text.lower()
    if not text or lower in NON_FILE_NIFTI_VALUES:
        return False
    if not any(lower.endswith(ext) for ext in NIFTI_EXTENSIONS):
        return False
    name = Path(text).name
    if name in NON_FILE_NIFTI_VALUES:
        return False
    # A suffix such as "_defaced.nii.gz" is not a complete file name.
    if "/" not in text and name.startswith("_"):
        return False
    return True


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def connect_source(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def open_output(path: Path, replace: bool = False) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    if replace and path.exists():
        path.unlink()
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    create_output_schema(conn)
    return conn


def create_output_schema(conn: sqlite3.Connection) -> None:
    normalized_sql = ",\n            ".join(f'"{name}" TEXT' for name in NORMALIZED_COLUMNS)
    conn.executescript(
        f"""
        CREATE TABLE IF NOT EXISTS harvest_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS nifti_downloads (
            download_row_id INTEGER PRIMARY KEY,
            parent_source TEXT,
            dataset_type TEXT,
            short_title TEXT,
            title TEXT,
            download_id TEXT,
            download_title TEXT,
            download_url TEXT,
            download_size TEXT,
            download_size_unit TEXT,
            subjects TEXT,
            studies TEXT,
            series TEXT,
            images TEXT,
            download_types TEXT,
            data_types TEXT,
            file_types TEXT,
            license_label TEXT,
            access_level TEXT,
            raw_json TEXT
        );

        CREATE TABLE IF NOT EXISTS candidate_downloads (
            download_row_id INTEGER PRIMARY KEY,
            short_title TEXT,
            dataset_type TEXT,
            download_id TEXT,
            download_title TEXT,
            download_url TEXT,
            candidate_kind TEXT,
            route TEXT,
            file_types TEXT,
            data_types TEXT,
            download_types TEXT,
            license_label TEXT,
            reason TEXT
        );

        CREATE TABLE IF NOT EXISTS package_files (
            package_file_id INTEGER PRIMARY KEY,
            download_row_id INTEGER,
            short_title TEXT,
            download_id TEXT,
            download_title TEXT,
            source_url TEXT,
            package_path TEXT,
            file_name TEXT,
            file_ext TEXT,
            bytes TEXT,
            row_json TEXT,
            is_metadata_candidate INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS harvested_files (
            file_id INTEGER PRIMARY KEY,
            download_row_id INTEGER,
            short_title TEXT,
            dataset_type TEXT,
            download_id TEXT,
            download_title TEXT,
            source_kind TEXT,
            source_url TEXT,
            package_path TEXT,
            local_path TEXT,
            file_name TEXT,
            file_ext TEXT,
            file_size_bytes INTEGER,
            sha256 TEXT,
            mime_type TEXT,
            status TEXT,
            error TEXT
        );

        CREATE TABLE IF NOT EXISTS tabular_sheets (
            sheet_id INTEGER PRIMARY KEY,
            file_id INTEGER,
            short_title TEXT,
            file_name TEXT,
            sheet_name TEXT,
            row_count INTEGER,
            column_count INTEGER,
            columns_json TEXT,
            parsed_status TEXT,
            error TEXT
        );

        CREATE TABLE IF NOT EXISTS tabular_rows (
            sheet_id INTEGER,
            file_id INTEGER,
            short_title TEXT,
            sheet_name TEXT,
            row_number INTEGER,
            row_json TEXT,
            PRIMARY KEY (sheet_id, row_number)
        );

        CREATE TABLE IF NOT EXISTS normalized_series_rows (
            normalized_row_id INTEGER PRIMARY KEY,
            file_id INTEGER,
            sheet_id INTEGER,
            short_title TEXT,
            dataset_type TEXT,
            download_id TEXT,
            download_title TEXT,
            source_kind TEXT,
            source_file_name TEXT,
            package_path TEXT,
            sheet_name TEXT,
            source_row_number INTEGER,
            {normalized_sql},
            matched_columns_json TEXT,
            raw_row_json TEXT
        );

        CREATE TABLE IF NOT EXISTS aspera_root_sums_inventory (
            sums_inventory_id INTEGER PRIMARY KEY,
            short_title TEXT,
            dataset_type TEXT,
            download_id TEXT,
            download_title TEXT,
            sums_package_path TEXT,
            local_sums_path TEXT,
            line_number INTEGER,
            checksum TEXT,
            algorithm TEXT,
            package_path TEXT,
            file_name TEXT,
            file_ext TEXT,
            raw_line TEXT
        );

        CREATE TABLE IF NOT EXISTS metadata_quality_flags (
            flag_id INTEGER PRIMARY KEY,
            short_title TEXT,
            dataset_type TEXT,
            affected_table TEXT,
            affected_row_id INTEGER,
            source_file_name TEXT,
            sheet_name TEXT,
            source_row_number INTEGER,
            excel_row INTEGER,
            nifti_file TEXT,
            related_download_id TEXT,
            issue_type TEXT,
            severity TEXT,
            status TEXT,
            evidence TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS nifti_file_series (
            nifti_file_series_id INTEGER PRIMARY KEY,
            short_title TEXT,
            dataset_type TEXT,
            download_ids TEXT,
            download_titles TEXT,
            file_name TEXT,
            file_ext TEXT,
            package_path TEXT,
            inventory_sources TEXT,
            metadata_sources TEXT,
            source_row_count INTEGER,
            checksum TEXT,
            checksum_algorithm TEXT,
            bytes TEXT,
            quality_flag_json TEXT,
            {normalized_sql}
        );

        CREATE TABLE IF NOT EXISTS non_dicom_files (
            non_dicom_file_id TEXT PRIMARY KEY,
            short_title TEXT,
            dataset_type TEXT,
            download_ids TEXT,
            download_titles TEXT,
            file_name TEXT,
            file_ext TEXT,
            package_path TEXT,
            file_group_id TEXT,
            file_role TEXT,
            inventory_sources TEXT,
            metadata_sources TEXT,
            source_table TEXT,
            source_row_id INTEGER,
            bytes TEXT,
            checksum TEXT,
            checksum_algorithm TEXT,
            is_nifti INTEGER NOT NULL DEFAULT 0,
            is_sidecar INTEGER NOT NULL DEFAULT 0,
            is_package_metadata INTEGER NOT NULL DEFAULT 0,
            is_derived_candidate INTEGER NOT NULL DEFAULT 0,
            quality_flag_json TEXT
        );

        CREATE TABLE IF NOT EXISTS radiology_series (
            radiology_id TEXT PRIMARY KEY,
            non_dicom_file_id TEXT,
            file_group_id TEXT,
            short_title TEXT,
            dataset_type TEXT,
            download_ids TEXT,
            file_name TEXT,
            package_path TEXT,
            subject_id TEXT,
            procedure_id TEXT,
            study_id TEXT,
            study_id_source TEXT,
            series_id TEXT,
            series_id_source TEXT,
            source_doi TEXT,
            modality TEXT,
            body_part_examined TEXT,
            study_date TEXT,
            series_date TEXT,
            study_description TEXT,
            series_description TEXT,
            series_number TEXT,
            manufacturer TEXT,
            manufacturer_model_name TEXT,
            software_versions TEXT,
            image_type TEXT,
            object_type TEXT,
            rows TEXT,
            columns TEXT,
            number_of_slices TEXT,
            number_of_temporal_positions TEXT,
            pixel_spacing_row_mm TEXT,
            pixel_spacing_col_mm TEXT,
            slice_thickness_mm TEXT,
            spacing_between_slices_mm TEXT,
            orientation_or_affine TEXT,
            is_phantom TEXT,
            is_derived_object INTEGER NOT NULL DEFAULT 0,
            quality_flag_json TEXT
        );

        CREATE TABLE IF NOT EXISTS radiology_mr (
            radiology_id TEXT PRIMARY KEY,
            magnetic_field_strength_t TEXT,
            scanning_sequence TEXT,
            sequence_variant TEXT,
            mr_acquisition_type TEXT,
            echo_time_ms TEXT,
            repetition_time_ms TEXT,
            echo_train_length TEXT,
            flip_angle_deg TEXT,
            pixel_bandwidth_hz TEXT,
            imaging_frequency_mhz TEXT,
            imaged_nucleus TEXT,
            inversion_time_ms TEXT,
            receive_coil_name TEXT,
            sequence_name TEXT,
            diffusion_b_value_s_per_mm2 TEXT
        );

        CREATE TABLE IF NOT EXISTS radiology_ct (
            radiology_id TEXT PRIMARY KEY,
            kvp TEXT,
            scan_options TEXT,
            convolution_kernel TEXT,
            gantry_detector_tilt_deg TEXT,
            xray_tube_current_min_ma TEXT,
            xray_tube_current_max_ma TEXT,
            filter_type TEXT,
            exposure_min_mas TEXT,
            exposure_max_mas TEXT,
            exposure_time_min_ms TEXT,
            exposure_time_max_ms TEXT,
            data_collection_diameter_mm TEXT,
            reconstruction_diameter_mm TEXT,
            spiral_pitch_factor TEXT
        );

        CREATE TABLE IF NOT EXISTS radiology_pet (
            radiology_id TEXT PRIMARY KEY,
            series_type TEXT,
            units TEXT,
            decay_correction TEXT,
            corrected_image TEXT,
            randoms_correction_method TEXT,
            reconstruction_method TEXT,
            actual_frame_duration_ms TEXT,
            scatter_correction_method TEXT,
            attenuation_correction_method TEXT,
            radionuclide_code_meaning TEXT,
            radionuclide_total_dose_bq TEXT,
            radiopharmaceutical_start_time TEXT,
            radiopharmaceutical TEXT,
            number_of_time_slices TEXT
        );

        CREATE TABLE IF NOT EXISTS radiology_contrast (
            radiology_id TEXT PRIMARY KEY,
            contrast_bolus_agent TEXT,
            contrast_bolus_ingredient TEXT,
            contrast_bolus_route TEXT
        );

        CREATE TABLE IF NOT EXISTS derived_objects (
            derived_object_id TEXT PRIMARY KEY,
            non_dicom_file_id TEXT,
            radiology_id TEXT,
            short_title TEXT,
            dataset_type TEXT,
            file_name TEXT,
            package_path TEXT,
            file_ext TEXT,
            analysis_result_id TEXT,
            referenced_radiology_id TEXT,
            referenced_series_id TEXT,
            derived_object_type TEXT,
            segmentation_representation TEXT,
            segmentation_type TEXT,
            total_segments TEXT,
            algorithm_type TEXT,
            algorithm_name TEXT,
            segmented_property_category TEXT,
            segmented_property_type TEXT,
            anatomic_region TEXT,
            roi_names TEXT,
            roi_generation_algorithms TEXT,
            rt_roi_interpreted_types TEXT,
            annotation_coordinate_type TEXT,
            derivation_basis TEXT
        );

        CREATE TABLE IF NOT EXISTS derived_object_references (
            derived_object_reference_id TEXT PRIMARY KEY,
            derived_object_id TEXT,
            derived_non_dicom_file_id TEXT,
            derived_radiology_id TEXT,
            referenced_non_dicom_file_id TEXT,
            referenced_radiology_id TEXT,
            referenced_series_id TEXT,
            referenced_file_name TEXT,
            referenced_package_path TEXT,
            reference_role TEXT,
            inference_method TEXT,
            confidence TEXT,
            evidence_json TEXT
        );

        CREATE TABLE IF NOT EXISTS annotation_groups (
            annotation_group_id TEXT PRIMARY KEY,
            derived_object_id TEXT,
            annotation_group_number TEXT,
            annotation_group_uid TEXT,
            annotation_group_label TEXT,
            annotation_group_generation_type TEXT,
            number_of_annotations TEXT,
            graphic_type TEXT,
            annotation_property_category_code TEXT,
            annotation_property_category_meaning TEXT,
            annotation_property_type_code TEXT,
            annotation_property_type_meaning TEXT,
            algorithm_name TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_candidate_downloads_short_title
          ON candidate_downloads(short_title);
        CREATE INDEX IF NOT EXISTS idx_package_files_download
          ON package_files(download_row_id);
        CREATE INDEX IF NOT EXISTS idx_harvested_files_short_title
          ON harvested_files(short_title);
        CREATE INDEX IF NOT EXISTS idx_normalized_short_title
          ON normalized_series_rows(short_title);
        CREATE INDEX IF NOT EXISTS idx_normalized_patient
          ON normalized_series_rows(PatientID);
        CREATE INDEX IF NOT EXISTS idx_normalized_series_uid
          ON normalized_series_rows(SeriesInstanceUID);
        CREATE INDEX IF NOT EXISTS idx_aspera_root_sums_dataset
          ON aspera_root_sums_inventory(short_title, download_id);
        CREATE INDEX IF NOT EXISTS idx_aspera_root_sums_ext
          ON aspera_root_sums_inventory(file_ext);
        CREATE INDEX IF NOT EXISTS idx_quality_flags_lookup
          ON metadata_quality_flags(short_title, affected_table, affected_row_id, issue_type);
        CREATE INDEX IF NOT EXISTS idx_nifti_file_series_dataset
          ON nifti_file_series(short_title);
        CREATE INDEX IF NOT EXISTS idx_nifti_file_series_file
          ON nifti_file_series(nifti_file);
        CREATE INDEX IF NOT EXISTS idx_non_dicom_files_dataset
          ON non_dicom_files(short_title);
        CREATE INDEX IF NOT EXISTS idx_non_dicom_files_group
          ON non_dicom_files(file_group_id);
        CREATE INDEX IF NOT EXISTS idx_radiology_series_dataset
          ON radiology_series(short_title);
        CREATE INDEX IF NOT EXISTS idx_radiology_series_file
          ON radiology_series(non_dicom_file_id);
        CREATE INDEX IF NOT EXISTS idx_radiology_series_modality
          ON radiology_series(modality);
        CREATE INDEX IF NOT EXISTS idx_radiology_series_series_id
          ON radiology_series(series_id);
        CREATE INDEX IF NOT EXISTS idx_derived_objects_dataset
          ON derived_objects(short_title);
        CREATE INDEX IF NOT EXISTS idx_derived_objects_file
          ON derived_objects(non_dicom_file_id);
        CREATE INDEX IF NOT EXISTS idx_derived_object_refs_object
          ON derived_object_references(derived_object_id);
        CREATE INDEX IF NOT EXISTS idx_derived_object_refs_source
          ON derived_object_references(referenced_non_dicom_file_id);

        CREATE VIEW IF NOT EXISTS agent_nifti_downloads AS
        SELECT
            download_row_id,
            parent_source,
            dataset_type,
            short_title,
            title,
            download_id,
            COALESCE(
                NULLIF(download_title, ''),
                trim(
                    short_title || ' ' ||
                    CASE
                        WHEN COALESCE(download_id, '') <> '' THEN 'download ' || download_id || ' '
                        ELSE ''
                    END ||
                    CASE
                        WHEN COALESCE(file_types, '') <> '' THEN file_types
                        ELSE ''
                    END
                )
            ) AS download_label,
            download_title,
            download_url,
            download_size,
            download_size_unit,
            subjects,
            studies,
            series,
            images,
            download_types,
            data_types,
            file_types,
            license_label,
            access_level
        FROM nifti_downloads;

        CREATE VIEW IF NOT EXISTS agent_nifti_dataset_summary AS
        WITH download_summary AS (
            SELECT
                parent_source,
                dataset_type,
                short_title,
                title,
                COUNT(*) AS nifti_downloads,
                group_concat(download_id, '; ') AS download_ids,
                group_concat(download_label, '; ') AS download_labels
            FROM agent_nifti_downloads
            GROUP BY parent_source, dataset_type, short_title, title
        ),
        file_summary AS (
            SELECT
                short_title,
                COUNT(*) AS non_dicom_files,
                SUM(CASE WHEN is_nifti THEN 1 ELSE 0 END) AS nifti_files,
                SUM(CASE WHEN is_sidecar THEN 1 ELSE 0 END) AS sidecar_files,
                SUM(CASE WHEN is_package_metadata THEN 1 ELSE 0 END) AS package_metadata_files
            FROM non_dicom_files
            GROUP BY short_title
        ),
        radiology_summary AS (
            SELECT
                short_title,
                COUNT(*) AS radiology_series_rows,
                SUM(CASE WHEN modality = 'MR' THEN 1 ELSE 0 END) AS mr_files,
                SUM(CASE WHEN modality = 'CT' THEN 1 ELSE 0 END) AS ct_files,
                SUM(CASE WHEN is_derived_object THEN 1 ELSE 0 END) AS derived_radiology_rows,
                COUNT(DISTINCT NULLIF(subject_id, '')) AS subject_ids,
                COUNT(DISTINCT NULLIF(study_id, '')) AS study_ids,
                COUNT(DISTINCT NULLIF(series_id, '')) AS series_ids
            FROM radiology_series
            GROUP BY short_title
        ),
        derived_summary AS (
            SELECT
                d.short_title,
                COUNT(*) AS derived_objects,
                COUNT(DISTINCT dor.derived_object_id) AS linked_derived_objects
            FROM derived_objects d
            LEFT JOIN derived_object_references dor
              ON dor.derived_object_id = d.derived_object_id
            GROUP BY d.short_title
        )
        SELECT
            d.parent_source,
            d.dataset_type,
            d.short_title,
            d.title,
            d.nifti_downloads,
            COALESCE(f.nifti_files, 0) AS nifti_files,
            COALESCE(f.non_dicom_files, 0) AS non_dicom_files,
            COALESCE(f.sidecar_files, 0) AS sidecar_files,
            COALESCE(f.package_metadata_files, 0) AS package_metadata_files,
            COALESCE(r.radiology_series_rows, 0) AS radiology_series_rows,
            COALESCE(r.mr_files, 0) AS mr_files,
            COALESCE(r.ct_files, 0) AS ct_files,
            COALESCE(r.derived_radiology_rows, 0) AS derived_radiology_rows,
            COALESCE(x.derived_objects, 0) AS derived_objects,
            COALESCE(x.linked_derived_objects, 0) AS linked_derived_objects,
            COALESCE(r.subject_ids, 0) AS subject_ids,
            COALESCE(r.study_ids, 0) AS study_ids,
            COALESCE(r.series_ids, 0) AS series_ids,
            d.download_ids,
            d.download_labels
        FROM download_summary d
        LEFT JOIN file_summary f ON lower(f.short_title) = lower(d.short_title)
        LEFT JOIN radiology_summary r ON lower(r.short_title) = lower(d.short_title)
        LEFT JOIN derived_summary x ON lower(x.short_title) = lower(d.short_title)
        ORDER BY lower(d.short_title);

        CREATE VIEW IF NOT EXISTS agent_nifti_files AS
        SELECT
            radiology_id,
            non_dicom_file_id,
            file_group_id,
            short_title,
            dataset_type,
            download_ids,
            download_ids AS download_id,
            file_name,
            package_path,
            subject_id,
            procedure_id,
            study_id,
            CASE WHEN study_id_source = 'source_metadata' THEN study_id ELSE '' END AS study_instance_uid,
            study_id_source,
            series_id,
            CASE WHEN series_id_source = 'source_metadata' THEN series_id ELSE '' END AS series_instance_uid,
            series_id_source,
            source_doi,
            modality,
            body_part_examined,
            study_date,
            series_date,
            study_description,
            series_description,
            series_number,
            manufacturer,
            manufacturer_model_name,
            software_versions,
            image_type,
            object_type,
            rows,
            columns,
            number_of_slices,
            number_of_temporal_positions,
            pixel_spacing_row_mm,
            pixel_spacing_col_mm,
            slice_thickness_mm,
            spacing_between_slices_mm,
            orientation_or_affine,
            is_phantom,
            is_derived_object,
            quality_flag_json
        FROM radiology_series;

        CREATE VIEW IF NOT EXISTS agent_nifti_derived_objects AS
        SELECT
            d.derived_object_id,
            d.non_dicom_file_id,
            d.radiology_id,
            d.short_title,
            d.dataset_type,
            d.file_name,
            d.package_path,
            d.file_ext,
            d.referenced_radiology_id,
            d.referenced_series_id,
            d.derived_object_type,
            d.segmentation_representation,
            d.segmentation_type,
            d.total_segments,
            d.algorithm_type,
            d.algorithm_name,
            dor.referenced_non_dicom_file_id,
            dor.referenced_file_name,
            dor.referenced_package_path,
            dor.reference_role,
            dor.inference_method,
            dor.confidence,
            dor.evidence_json
        FROM derived_objects d
        LEFT JOIN derived_object_references dor
          ON dor.derived_object_id = d.derived_object_id;
        """
    )
    ensure_column(conn, "derived_objects", "segmentation_representation", "TEXT")
    conn.commit()


def ensure_column(
    conn: sqlite3.Connection, table_name: str, column_name: str, column_definition: str
) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})")}
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_definition}")


def route_for_url(url: str) -> str:
    lower = url.lower()
    if not url:
        return "blank"
    if "faspex" in lower or "aspera" in lower:
        return "aspera"
    ext = extension_for_name(lower)
    if ext in ZIP_EXTENSIONS:
        return "direct_zip"
    if ext in SPREADSHEET_EXTENSIONS:
        return "direct_spreadsheet"
    return "direct_other"


def has_label(labels_json: str, label: str) -> bool:
    try:
        labels = json.loads(labels_json or "[]")
    except json.JSONDecodeError:
        return False
    return any(str(item).lower() == label.lower() for item in labels)


def labels_contain_any(labels_json: str, wanted: Iterable[str]) -> bool:
    wanted_norm = {item.lower() for item in wanted}
    try:
        labels = json.loads(labels_json or "[]")
    except json.JSONDecodeError:
        return False
    return any(str(item).lower() in wanted_norm for item in labels)


def text_has_metadata_hint(*values: str) -> bool:
    text = " ".join(value or "" for value in values).lower()
    return any(hint in text for hint in METADATA_NAME_HINTS)


def text_has_skip_hint(*values: str) -> bool:
    text = " ".join(value or "" for value in values).lower()
    return any(hint in text for hint in SKIP_NAME_HINTS)


def load_nifti_downloads(source: sqlite3.Connection) -> list[sqlite3.Row]:
    sql = """
        SELECT DISTINCT d.*
        FROM agent_current_downloads d
        JOIN wordpress_download_labels l
          ON l.download_row_id = d.download_row_id
        WHERE d.hidden = 0
          AND d.controlled_access = 0
          AND l.label_kind = 'file_type'
          AND lower(l.label) = 'nifti'
        ORDER BY lower(d.short_title), d.download_id, d.download_title
    """
    return source.execute(sql).fetchall()


def load_candidate_downloads(
    source: sqlite3.Connection, short_titles: list[str]
) -> list[dict[str, Any]]:
    if not short_titles:
        return []
    placeholders = ",".join("?" for _ in short_titles)
    sql = f"""
        SELECT d.*
        FROM agent_current_downloads d
        WHERE d.hidden = 0
          AND d.controlled_access = 0
          AND d.short_title IN ({placeholders})
        ORDER BY lower(d.short_title), d.download_id, d.download_title
    """
    candidates: list[dict[str, Any]] = []
    for row in source.execute(sql, short_titles).fetchall():
        file_types = row["file_types"] or "[]"
        route = route_for_url(row["download_url"] or "")
        is_nifti = has_label(file_types, "NIfTI")
        is_sheet = labels_contain_any(file_types, ("CSV", "TSV", "XLSX"))
        is_zip = labels_contain_any(file_types, ("ZIP",))
        hint = text_has_metadata_hint(
            row["download_title"] or "",
            row["description"] or "",
            row["download_url"] or "",
        )
        if not (is_nifti or is_sheet or is_zip or hint):
            continue
        if is_nifti and route == "aspera":
            kind = "nifti_aspera_package"
        elif is_nifti and route == "direct_zip":
            kind = "nifti_direct_zip"
        elif is_sheet:
            kind = "companion_spreadsheet"
        elif is_zip:
            kind = "companion_zip"
        else:
            kind = "metadata_hint"
        candidates.append(
            {
                **dict(row),
                "candidate_kind": kind,
                "route": route,
                "reason": ",".join(
                    part
                    for part, include in [
                        ("nifti", is_nifti),
                        ("sheet_label", is_sheet),
                        ("zip_label", is_zip),
                        ("metadata_hint", hint),
                    ]
                    if include
                ),
            }
        )
    return candidates


def insert_nifti_downloads(conn: sqlite3.Connection, rows: list[sqlite3.Row]) -> None:
    conn.executemany(
        """
        INSERT OR REPLACE INTO nifti_downloads
        (download_row_id, parent_source, dataset_type, short_title, title,
         download_id, download_title, download_url, download_size, download_size_unit,
         subjects, studies, series, images, download_types, data_types, file_types,
         license_label, access_level, raw_json)
        VALUES
        (:download_row_id, :parent_source, :dataset_type, :short_title, :title,
         :download_id, :download_title, :download_url, :download_size, :download_size_unit,
         :subjects, :studies, :series, :images, :download_types, :data_types, :file_types,
         :license_label, :access_level, :raw_json)
        """,
        [dict(row) for row in rows],
    )
    conn.commit()


def insert_candidate_downloads(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> None:
    conn.executemany(
        """
        INSERT OR REPLACE INTO candidate_downloads
        (download_row_id, short_title, dataset_type, download_id, download_title,
         download_url, candidate_kind, route, file_types, data_types, download_types,
         license_label, reason)
        VALUES
        (:download_row_id, :short_title, :dataset_type, :download_id, :download_title,
         :download_url, :candidate_kind, :route, :file_types, :data_types, :download_types,
         :license_label, :reason)
        """,
        rows,
    )
    conn.commit()


def request_bytes(url: str, timeout: int = 180) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read()


def download_url_to_file(url: str, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=180) as response:
        with out_path.open("wb") as target:
            shutil.copyfileobj(response, target)


def insert_harvested_file(
    conn: sqlite3.Connection,
    candidate: dict[str, Any],
    source_kind: str,
    local_path: Path | None,
    file_name: str,
    source_url: str = "",
    package_path: str = "",
    status: str = "ok",
    error: str = "",
) -> int:
    file_size = local_path.stat().st_size if local_path and local_path.exists() else None
    digest = sha256_file(local_path) if local_path and local_path.exists() else ""
    mime_type = mimetypes.guess_type(file_name)[0] or ""
    cur = conn.execute(
        """
        INSERT INTO harvested_files
        (download_row_id, short_title, dataset_type, download_id, download_title,
         source_kind, source_url, package_path, local_path, file_name, file_ext,
         file_size_bytes, sha256, mime_type, status, error)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            candidate.get("download_row_id"),
            candidate.get("short_title"),
            candidate.get("dataset_type"),
            candidate.get("download_id"),
            candidate.get("download_title"),
            source_kind,
            source_url or candidate.get("download_url", ""),
            package_path,
            str(local_path) if local_path else "",
            file_name,
            extension_for_name(file_name),
            file_size,
            digest,
            mime_type,
            status,
            error,
        ),
    )
    conn.commit()
    return int(cur.lastrowid)


def sniff_delimiter(sample: str, fallback: str = ",") -> str:
    try:
        return csv.Sniffer().sniff(sample[:8192], delimiters=",\t;|").delimiter
    except csv.Error:
        return fallback


def clean_header(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, 1):
        text = str(value).strip() if value is not None else ""
        if not text:
            text = f"column_{index}"
        count = used.get(text, 0)
        used[text] = count + 1
        if count:
            text = f"{text}_{count + 1}"
        headers.append(text)
    return headers


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def parse_csv_like(path: Path, delimiter: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    delimiter = delimiter or sniff_delimiter(text, "\t" if path.suffix.lower() == ".tsv" else ",")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        headers = clean_header(next(reader))
    except StopIteration:
        return [], []
    rows = [dict(zip(headers, row)) for row in reader]
    return headers, rows


def parse_xlsx(path: Path) -> list[tuple[str, list[str], list[dict[str, str]]]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse XLSX files") from exc
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parsed: list[tuple[str, list[str], list[dict[str, str]]]] = []
    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        header_values: list[Any] | None = None
        for values in rows_iter:
            if values and any(value is not None and str(value).strip() for value in values):
                header_values = list(values)
                break
        if header_values is None:
            parsed.append((worksheet.title, [], []))
            continue
        headers = clean_header(header_values)
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            if not values or not any(value is not None and str(value).strip() for value in values):
                continue
            padded = list(values) + [""] * max(0, len(headers) - len(values))
            rows.append(
                {header: stringify_cell(padded[index]) for index, header in enumerate(headers)}
            )
        parsed.append((worksheet.title, headers, rows))
    workbook.close()
    return parsed


def column_lookup(headers: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for header in headers:
        normalized = normalize_key(header)
        if normalized and normalized not in lookup:
            lookup[normalized] = header
    return lookup


def extract_normalized(
    headers: list[str], row: dict[str, str]
) -> tuple[dict[str, str], dict[str, str]]:
    lookup = column_lookup(headers)
    extracted: dict[str, str] = {}
    matched: dict[str, str] = {}
    for field, synonyms in FIELD_SYNONYMS.items():
        for synonym in synonyms:
            key = normalize_key(synonym)
            header = lookup.get(key)
            if header is None:
                continue
            value = str(row.get(header, "")).strip()
            if value:
                if field == "nifti_file" and not is_plausible_nifti_file(value):
                    continue
                extracted[field] = clean_package_path(value) if field == "nifti_file" else value
                matched[field] = header
                break
    if "nifti_file" not in extracted:
        for header in headers:
            value = str(row.get(header, "")).strip()
            if is_plausible_nifti_file(value):
                extracted["nifti_file"] = clean_package_path(value)
                matched["nifti_file"] = header
                break
    return extracted, matched


def row_has_metadata(extracted: dict[str, str]) -> bool:
    useful = set(extracted) - {"nifti_file"}
    return bool(useful) or "nifti_file" in extracted


def insert_sheet_and_rows(
    conn: sqlite3.Connection,
    file_id: int,
    file_meta: sqlite3.Row,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, str]],
    max_rows: int | None,
) -> None:
    if max_rows is not None:
        rows = rows[:max_rows]
    cur = conn.execute(
        """
        INSERT INTO tabular_sheets
        (file_id, short_title, file_name, sheet_name, row_count, column_count,
         columns_json, parsed_status, error)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            file_id,
            file_meta["short_title"],
            file_meta["file_name"],
            sheet_name,
            len(rows),
            len(headers),
            json_dumps(headers),
            "ok",
            "",
        ),
    )
    sheet_id = int(cur.lastrowid)
    row_records = []
    normalized_records = []
    for row_number, row in enumerate(rows, 1):
        row_json = json_dumps(row)
        row_records.append(
            (
                sheet_id,
                file_id,
                file_meta["short_title"],
                sheet_name,
                row_number,
                row_json,
            )
        )
        extracted, matched = extract_normalized(headers, row)
        if not row_has_metadata(extracted):
            continue
        normalized_values = {name: "" for name in NORMALIZED_COLUMNS}
        normalized_values.update(extracted)
        normalized_values["collection_id"] = normalized_values["collection_id"] or file_meta[
            "short_title"
        ]
        normalized_values["license_short_name"] = normalized_values["license_short_name"] or ""
        normalized_records.append(
            (
                file_id,
                sheet_id,
                file_meta["short_title"],
                file_meta["dataset_type"],
                file_meta["download_id"],
                file_meta["download_title"],
                file_meta["source_kind"],
                file_meta["file_name"],
                file_meta["package_path"],
                sheet_name,
                row_number,
                *[normalized_values[name] for name in NORMALIZED_COLUMNS],
                json_dumps(matched),
                row_json,
            )
        )
    if row_records:
        conn.executemany(
            """
            INSERT OR REPLACE INTO tabular_rows
            (sheet_id, file_id, short_title, sheet_name, row_number, row_json)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            row_records,
        )
    if normalized_records:
        placeholders = ",".join("?" for _ in range(11 + len(NORMALIZED_COLUMNS) + 2))
        quoted_columns = ", ".join(f'"{name}"' for name in NORMALIZED_COLUMNS)
        conn.executemany(
            f"""
            INSERT INTO normalized_series_rows
            (file_id, sheet_id, short_title, dataset_type, download_id, download_title,
             source_kind, source_file_name, package_path, sheet_name, source_row_number,
             {quoted_columns}, matched_columns_json, raw_row_json)
            VALUES ({placeholders})
            """,
            normalized_records,
        )
    conn.commit()


def parse_harvested_file(
    conn: sqlite3.Connection, file_id: int, max_rows: int | None = None
) -> None:
    file_meta = conn.execute(
        "SELECT * FROM harvested_files WHERE file_id = ?", (file_id,)
    ).fetchone()
    if not file_meta or file_meta["status"] != "ok":
        return
    if str(file_meta["source_kind"] or "").startswith("aspera_") and str(
        file_meta["source_kind"] or ""
    ).endswith("_listing"):
        return
    path = Path(file_meta["local_path"])
    ext = extension_for_name(file_meta["file_name"])
    try:
        if ext == ".csv":
            headers, rows = parse_csv_like(path, ",")
            insert_sheet_and_rows(conn, file_id, file_meta, "csv", headers, rows, max_rows)
        elif ext == ".tsv":
            headers, rows = parse_csv_like(path, "\t")
            insert_sheet_and_rows(conn, file_id, file_meta, "tsv", headers, rows, max_rows)
        elif ext == ".xlsx":
            for sheet_name, headers, rows in parse_xlsx(path):
                insert_sheet_and_rows(conn, file_id, file_meta, sheet_name, headers, rows, max_rows)
    except Exception as exc:  # noqa: BLE001 - record parse failures in the output DB.
        conn.execute(
            """
            INSERT INTO tabular_sheets
            (file_id, short_title, file_name, sheet_name, row_count, column_count,
             columns_json, parsed_status, error)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                file_id,
                file_meta["short_title"],
                file_meta["file_name"],
                "",
                0,
                0,
                "[]",
                "error",
                str(exc),
            ),
        )
        conn.commit()


def candidate_local_path(base_dir: Path, candidate: dict[str, Any], file_name: str) -> Path:
    return (
        base_dir
        / safe_slug(candidate.get("short_title") or "dataset")
        / safe_slug(str(candidate.get("download_id") or candidate.get("download_row_id")))
        / safe_slug(file_name, "download")
    )


def harvest_direct_candidate(
    conn: sqlite3.Connection,
    candidate: dict[str, Any],
    files_dir: Path,
    max_zip_member_bytes: int,
    max_rows: int | None,
    reuse_cache: bool,
    cache_only: bool,
) -> None:
    url = candidate.get("download_url") or ""
    if not url:
        return
    route = candidate.get("route") or route_for_url(url)
    parsed_url = urllib.parse.urlparse(url)
    name = Path(parsed_url.path).name or f"download_{candidate['download_row_id']}"
    ext = extension_for_name(name)
    if route == "direct_spreadsheet" or ext in SPREADSHEET_EXTENSIONS:
        local_path = candidate_local_path(files_dir, candidate, name)
        try:
            if not (reuse_cache and local_path.exists()):
                if cache_only:
                    raise FileNotFoundError(f"cached file not found: {local_path}")
                download_url_to_file(url, local_path)
            file_id = insert_harvested_file(
                conn, candidate, "direct_download", local_path, name, source_url=url
            )
            parse_harvested_file(conn, file_id, max_rows)
        except Exception as exc:  # noqa: BLE001
            insert_harvested_file(
                conn,
                candidate,
                "direct_download",
                None,
                name,
                source_url=url,
                status="error",
                error=str(exc),
            )
    elif route == "direct_zip" or ext == ".zip":
        local_zip = candidate_local_path(files_dir, candidate, name)
        try:
            if not (reuse_cache and local_zip.exists()):
                if cache_only:
                    raise FileNotFoundError(f"cached file not found: {local_zip}")
                download_url_to_file(url, local_zip)
            insert_harvested_file(
                conn, candidate, "direct_zip", local_zip, name, source_url=url
            )
            harvest_zip_members(
                conn, candidate, local_zip, files_dir, max_zip_member_bytes, max_rows
            )
        except Exception as exc:  # noqa: BLE001
            insert_harvested_file(
                conn,
                candidate,
                "direct_zip",
                None,
                name,
                source_url=url,
                status="error",
                error=str(exc),
            )


def zip_member_is_metadata(name: str) -> bool:
    ext = extension_for_name(name)
    if ext in SPREADSHEET_EXTENSIONS:
        return True
    lower = name.lower()
    if text_has_skip_hint(lower):
        return False
    return text_has_metadata_hint(lower) and ext in {".txt", ".json"}


def zip_member_is_inventory_file(name: str) -> bool:
    parts = split_package_path(name)
    if not parts:
        return False
    if any(part == "__MACOSX" or part.startswith("._") for part in parts):
        return False
    return is_plausible_nifti_file(name)


def insert_direct_zip_nifti_package_files(
    conn: sqlite3.Connection, candidate: dict[str, Any], zip_path: Path
) -> int:
    existing = {
        row["package_path"]
        for row in conn.execute(
            """
            SELECT package_path
            FROM package_files
            WHERE short_title = ?
              AND download_id = ?
            """,
            (candidate.get("short_title"), candidate.get("download_id")),
        )
    }
    records = []
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            package_path = clean_package_path(info.filename)
            if info.is_dir() or not zip_member_is_inventory_file(package_path):
                continue
            if package_path in existing:
                continue
            row_json = {
                "source_kind": "direct_zip_member",
                "zip_file": str(zip_path),
                "filename": info.filename,
                "file_size": info.file_size,
                "compress_size": info.compress_size,
                "crc": info.CRC,
            }
            records.append(
                (
                    candidate.get("download_row_id"),
                    candidate.get("short_title"),
                    candidate.get("download_id"),
                    candidate.get("download_title"),
                    str(zip_path),
                    package_path,
                    Path(package_path).name,
                    extension_for_name(package_path),
                    str(info.file_size),
                    json_dumps(row_json),
                    0,
                )
            )
    if records:
        conn.executemany(
            """
            INSERT INTO package_files
            (download_row_id, short_title, download_id, download_title, source_url,
             package_path, file_name, file_ext, bytes, row_json, is_metadata_candidate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            records,
        )
        conn.commit()
    return len(records)


def local_extract_search_bases(files_dir: Path) -> list[Path]:
    bases: list[Path] = []
    env_value = os.environ.get("TCIA_NIFTI_LOCAL_EXTRACT_ROOTS", "")
    for token in env_value.split(os.pathsep):
        token = token.strip()
        if token:
            bases.append(Path(token).expanduser())
    for base in (Path.cwd(), files_dir, SKILL_ROOT):
        bases.append(base)
        bases.extend(list(base.parents)[:4])
    unique: list[Path] = []
    seen: set[str] = set()
    for base in bases:
        key = str(base)
        if key not in seen:
            seen.add(key)
            unique.append(base)
    return unique


def local_support_search_bases(files_dir: Path) -> list[Path]:
    bases: list[Path] = []
    env_value = os.environ.get("TCIA_NIFTI_LOCAL_SUPPORT_ROOTS", "")
    for token in env_value.split(os.pathsep):
        token = token.strip()
        if token:
            bases.append(Path(token).expanduser())
    bases.extend(local_extract_search_bases(files_dir))
    unique: list[Path] = []
    seen: set[str] = set()
    for base in bases:
        key = str(base)
        if key not in seen:
            seen.add(key)
            unique.append(base)
    return unique


def find_local_extracted_package_dir(files_dir: Path, dirname: str) -> Path | None:
    for base in local_extract_search_bases(files_dir):
        candidates = [base] if base.name == dirname else [base / dirname]
        for candidate in candidates:
            if candidate.is_dir():
                return candidate
    return None


def local_extract_member_is_inventory_file(path: Path) -> bool:
    parts = path.parts
    if any(part == "__MACOSX" or part.startswith("._") for part in parts):
        return False
    if path.name == ".DS_Store":
        return False
    return path.is_file()


def insert_local_extracted_package_files(
    conn: sqlite3.Connection, short_title: str, dirname: str, extracted_dir: Path
) -> int:
    candidate_row = conn.execute(
        """
        SELECT download_row_id, short_title, dataset_type, download_id,
               download_title, download_url
        FROM candidate_downloads
        WHERE short_title = ?
        ORDER BY download_row_id
        LIMIT 1
        """,
        (short_title,),
    ).fetchone()
    if candidate_row is None:
        return 0
    candidate = dict(candidate_row)
    existing = {
        row["package_path"]
        for row in conn.execute(
            """
            SELECT package_path
            FROM package_files
            WHERE short_title = ?
              AND download_id = ?
            """,
            (candidate.get("short_title"), candidate.get("download_id")),
        )
    }
    records = []
    for path in sorted(extracted_dir.rglob("*")):
        if not local_extract_member_is_inventory_file(path):
            continue
        relative_path = clean_package_path(str(path.relative_to(extracted_dir)))
        if not relative_path:
            continue
        package_path = clean_package_path(f"{short_title}/{dirname}/{relative_path}")
        if package_path in existing:
            continue
        row_json = {
            "source_kind": "local_extracted_package",
            "extracted_dir_name": dirname,
            "relative_path": relative_path,
            "file_size": path.stat().st_size,
        }
        records.append(
            (
                candidate.get("download_row_id"),
                candidate.get("short_title"),
                candidate.get("download_id"),
                candidate.get("download_title"),
                candidate.get("download_url"),
                package_path,
                Path(package_path).name,
                extension_for_name(package_path),
                str(path.stat().st_size),
                json_dumps(row_json),
                1 if aspera_file_is_metadata(package_path) else 0,
            )
        )
    if records:
        conn.executemany(
            """
            INSERT INTO package_files
            (download_row_id, short_title, download_id, download_title, source_url,
             package_path, file_name, file_ext, bytes, row_json, is_metadata_candidate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            records,
        )
        conn.commit()
    return len(records)


def inventory_local_extracted_package_files(conn: sqlite3.Connection, files_dir: Path) -> int:
    inserted = 0
    for spec in LOCAL_EXTRACTED_PACKAGE_SPECS:
        extracted_dir = find_local_extracted_package_dir(files_dir, spec["dirname"])
        if not extracted_dir:
            continue
        inserted += insert_local_extracted_package_files(
            conn,
            spec["short_title"],
            spec["dirname"],
            extracted_dir,
        )
    return inserted


def harvest_zip_members(
    conn: sqlite3.Connection,
    candidate: dict[str, Any],
    zip_path: Path,
    files_dir: Path,
    max_zip_member_bytes: int,
    max_rows: int | None,
) -> None:
    insert_direct_zip_nifti_package_files(conn, candidate, zip_path)
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            if info.is_dir() or not zip_member_is_metadata(info.filename):
                continue
            if info.file_size > max_zip_member_bytes:
                insert_harvested_file(
                    conn,
                    candidate,
                    "zip_member",
                    None,
                    Path(info.filename).name,
                    source_url=str(zip_path),
                    package_path=info.filename,
                    status="skipped",
                    error=f"zip member exceeds {max_zip_member_bytes} bytes",
                )
                continue
            out_path = candidate_local_path(files_dir, candidate, info.filename)
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(info) as source, out_path.open("wb") as target:
                shutil.copyfileobj(source, target)
            file_id = insert_harvested_file(
                conn,
                candidate,
                "zip_member",
                out_path,
                Path(info.filename).name,
                source_url=str(zip_path),
                package_path=info.filename,
            )
            parse_harvested_file(conn, file_id, max_rows)


def parse_ascli_csv(stdout: str) -> list[dict[str, str]]:
    lines = [
        line
        for line in stdout.splitlines()
        if line.strip() and not line.strip().startswith("Items:")
    ]
    if not lines:
        return []
    rows = list(csv.reader(lines))
    header = [value.strip().lower() for value in rows[0]]
    known_headers = {"path", "name", "type", "entry_type", "bytes", "size", "modified_at"}
    entry_types = {"directory", "dir", "folder", "file", "symbolic_link", "symlink"}

    def inferred_record(values: list[str]) -> dict[str, str]:
        path_candidates = [
            value
            for value in values
            if "/" in value
            and "view, edit, delete" not in value.lower()
            and not re.match(r"^\d{4}-\d{2}-\d{2}T", value)
        ]
        path = next(
            (value for value in path_candidates if "." in Path(value).name),
            next(iter(path_candidates), ""),
        )
        entry_type = next(
            (value for value in values if value.lower() in entry_types),
            "",
        )
        bytes_value = next((value for value in values if value.isdigit()), "")
        modified_at = next(
            (value for value in values if re.match(r"^\d{4}-\d{2}-\d{2}T", value)),
            "",
        )
        permissions = next((value for value in values if "view" in value.lower()), "")
        name = ""
        if path:
            path_name = Path(path).name
            name = next((value for value in values if value == path_name), path_name)
        if not name:
            name = next(
                (
                    value
                    for value in values
                    if value
                    and value.lower() not in entry_types
                    and value != bytes_value
                    and value != modified_at
                    and value != permissions
                    and "/" not in value
                ),
                "",
            )
        return {
            "path": clean_package_path(path),
            "name": name,
            "entry_type": entry_type,
            "bytes": bytes_value,
            "modified_at": modified_at,
            "permissions": permissions,
        }

    if known_headers.intersection(header):
        reader = csv.DictReader(io.StringIO(stdout))
        output: list[dict[str, str]] = []
        for row in reader:
            cleaned_row = {key: str(value or "").strip() for key, value in row.items()}
            inferred = inferred_record(
                [
                    value
                    for key, value in cleaned_row.items()
                    if key != "_browsed_path"
                ]
            )
            cleaned_row.update(inferred)
            output.append(cleaned_row)
        return output

    output: list[dict[str, str]] = []
    for row in rows:
        values = [str(value or "").strip() for value in row]
        record = inferred_record(values)
        record.update({f"raw_{index}": value for index, value in enumerate(values)})
        output.append(record)
    return output


def package_row_path(row: dict[str, str]) -> str:
    entry_type = package_row_entry_type(row)

    def path_like(value: str) -> bool:
        text = clean_package_path(value)
        if not text or "view, edit, delete" in text.lower():
            return False
        if re.match(r"^\d{4}-\d{2}-\d{2}T", text):
            return False
        if text.lower() in NON_FILE_NIFTI_VALUES:
            return False
        return (
            "/" in text
            or "." in Path(text).name
            or entry_type in {"directory", "dir", "folder"}
        )

    for key in ("path", "Path", "file", "File", "name", "Name"):
        value = row.get(key)
        if value and path_like(value):
            cleaned = clean_package_path(value)
            browsed_path = clean_package_path(row.get("_browsed_path", ""))
            if browsed_path and "/" not in cleaned and "." in Path(cleaned).name:
                return clean_package_path(f"{browsed_path}/{cleaned}")
            return cleaned
    for key, value in row.items():
        if value and path_like(value):
            cleaned = clean_package_path(value)
            browsed_path = clean_package_path(row.get("_browsed_path", ""))
            if browsed_path and "/" not in cleaned and "." in Path(cleaned).name:
                return clean_package_path(f"{browsed_path}/{cleaned}")
            return cleaned
    return ""


def package_row_entry_type(row: dict[str, str]) -> str:
    for key in ("type", "Type", "entry_type", "Entry Type", "directory"):
        value = row.get(key)
        if value:
            return str(value).strip().lower()
    return ""


def package_row_is_directory(row: dict[str, str]) -> bool:
    return package_row_entry_type(row) in {"directory", "dir", "folder"}


def package_row_bytes(row: dict[str, str]) -> str:
    for key in ("bytes", "size", "Size", "Bytes", "length", "Length"):
        value = row.get(key)
        if value and str(value).strip().isdigit():
            return value
    for value in row.values():
        text = str(value or "").strip()
        if text.isdigit():
            return text
    return ""


def aspera_file_is_metadata(path: str) -> bool:
    ext = extension_for_name(path)
    if ext in SPREADSHEET_EXTENSIONS:
        return True
    lower = path.lower()
    if text_has_skip_hint(lower):
        return False
    return text_has_metadata_hint(lower) and ext in {".txt", ".json"}


def browse_aspera_package(
    ascli: str,
    url: str,
    timeout: int,
    query: dict[str, Any] | None = None,
) -> list[dict[str, str]]:
    browse_query = {"recursive": True, **(query or {})}
    command = [
        ascli,
        "--format=csv",
        "faspex5",
        "packages",
        "browse",
        f"--query=@json:{json.dumps(browse_query, separators=(',', ':'))}",
        f"--url={url}",
    ]
    result = subprocess.run(command, check=True, capture_output=True, text=True, timeout=timeout)
    return parse_ascli_csv(result.stdout)


def browse_aspera_path(
    ascli: str,
    url: str,
    package_path: str,
    timeout: int,
    query: dict[str, Any] | None = None,
) -> list[dict[str, str]]:
    command = [
        ascli,
        "--format=csv",
        "faspex5",
        "packages",
        "browse",
        f"--url={url}",
    ]
    if query:
        command.append(f"--query=@json:{json.dumps(query, separators=(',', ':'))}")
    if package_path:
        command.append(package_path)
    result = subprocess.run(command, check=True, capture_output=True, text=True, timeout=timeout)
    return parse_ascli_csv(result.stdout)


def browse_aspera_interactive(
    ascli: str,
    url: str,
    timeout: int,
    max_dirs: int,
    max_depth: int,
    descend_all: bool,
) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    queue: list[tuple[str, int]] = [("", 0)]
    seen_dirs = {""}
    browsed_dirs = 0
    while queue and browsed_dirs < max_dirs:
        package_path, depth = queue.pop(0)
        try:
            rows = browse_aspera_path(ascli, url, package_path, timeout)
        except Exception as exc:  # noqa: BLE001 - keep fallback best-effort.
            output.append(
                {
                    "path": package_path,
                    "directory": "browse_error",
                    "error": str(exc),
                }
            )
            continue
        browsed_dirs += 1
        for row in rows:
            if package_path:
                row = {**row, "_browsed_path": package_path}
            output.append(row)
            path = package_row_path(row)
            if (
                not path
                or not package_row_is_directory(row)
                or path in seen_dirs
                or depth >= max_depth
            ):
                continue
            should_descend = descend_all or text_has_metadata_hint(path)
            if should_descend:
                seen_dirs.add(path)
                queue.append((path, depth + 1))
    return output


def insert_package_files(
    conn: sqlite3.Connection, candidate: dict[str, Any], rows: list[dict[str, str]]
) -> list[dict[str, str]]:
    metadata_rows: list[dict[str, str]] = []
    records = []
    for row in rows:
        path = package_row_path(row)
        if not path:
            continue
        ext = extension_for_name(path)
        is_metadata = aspera_file_is_metadata(path)
        if is_metadata:
            metadata_rows.append(row)
        records.append(
            (
                candidate.get("download_row_id"),
                candidate.get("short_title"),
                candidate.get("download_id"),
                candidate.get("download_title"),
                candidate.get("download_url"),
                path,
                Path(path).name,
                ext,
                package_row_bytes(row),
                json_dumps(row),
                1 if is_metadata else 0,
            )
        )
    if records:
        conn.executemany(
            """
            INSERT INTO package_files
            (download_row_id, short_title, download_id, download_title, source_url,
             package_path, file_name, file_ext, bytes, row_json, is_metadata_candidate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            records,
        )
        conn.commit()
    return metadata_rows


def receive_aspera_file(
    ascli: str, url: str, package_path: str, out_dir: Path, timeout: int
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    command = [
        ascli,
        "faspex5",
        "packages",
        "receive",
        f"--url={url}",
        package_path,
        f"--to-folder={out_dir}",
    ]
    subprocess.run(command, check=True, capture_output=True, text=True, timeout=timeout)
    candidates = sorted(
        (path for path in out_dir.rglob(Path(package_path).name) if path.is_file()),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"ascli did not produce expected file: {package_path}")
    return candidates[0]


def harvest_aspera_candidate(
    conn: sqlite3.Connection,
    candidate: dict[str, Any],
    files_dir: Path,
    ascli: str,
    max_metadata_files: int,
    max_rows: int | None,
    download_metadata_files: bool,
    timeout: int,
    interactive_fallback: bool,
    interactive_max_dirs: int,
    interactive_max_depth: int,
    interactive_descend_all: bool,
    reuse_cache: bool,
    cache_only: bool,
) -> None:
    url = candidate.get("download_url") or ""
    if not url:
        return
    existing = conn.execute(
        """
        SELECT 1
        FROM harvested_files
        WHERE download_row_id = ?
          AND source_kind = 'aspera_listing'
          AND status = 'ok'
        LIMIT 1
        """,
        (candidate.get("download_row_id"),),
    ).fetchone()
    if existing:
        return
    source_kind = "aspera_listing"
    listing_path = candidate_local_path(files_dir, candidate, "aspera_package_listing.csv")
    try:
        if reuse_cache and listing_path.exists():
            rows = parse_ascli_csv(listing_path.read_text(encoding="utf-8", errors="replace"))
        else:
            if cache_only:
                raise FileNotFoundError(f"cached Aspera listing not found: {listing_path}")
            rows = browse_aspera_package(ascli, url, timeout)
    except Exception as exc:  # noqa: BLE001
        if not interactive_fallback:
            insert_harvested_file(
                conn,
                candidate,
                "aspera_browse",
                None,
                "aspera_package_listing.csv",
                source_url=url,
                status="error",
                error=str(exc),
            )
            return
        source_kind = "aspera_interactive_listing"
        try:
            if cache_only:
                raise exc
            rows = browse_aspera_interactive(
                ascli,
                url,
                timeout,
                interactive_max_dirs,
                interactive_max_depth,
                interactive_descend_all,
            )
            rows.insert(
                0,
                {
                    "path": "",
                    "directory": "recursive_browse_error",
                    "error": str(exc),
                },
            )
        except Exception as fallback_exc:  # noqa: BLE001
            insert_harvested_file(
                conn,
                candidate,
                "aspera_browse",
                None,
                "aspera_package_listing.csv",
                source_url=url,
                status="error",
                error=f"recursive: {exc}; interactive: {fallback_exc}",
            )
            return
    metadata_rows = insert_package_files(conn, candidate, rows)
    listing_path.parent.mkdir(parents=True, exist_ok=True)
    if rows:
        with listing_path.open("w", newline="", encoding="utf-8") as stream:
            fieldnames = sorted({key for row in rows for key in row})
            writer = csv.DictWriter(stream, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    file_id = insert_harvested_file(
        conn,
        candidate,
        source_kind,
        listing_path,
        "aspera_package_listing.csv",
        source_url=url,
    )
    parse_harvested_file(conn, file_id, max_rows)
    if not download_metadata_files:
        return
    for row in metadata_rows[:max_metadata_files]:
        package_path = package_row_path(row)
        if not package_path:
            continue
        local_dir = (
            files_dir
            / safe_slug(candidate.get("short_title") or "dataset")
            / safe_slug(str(candidate.get("download_id") or candidate.get("download_row_id")))
            / "aspera_members"
        )
        try:
            local_path = receive_aspera_file(ascli, url, package_path, local_dir, timeout)
            file_id = insert_harvested_file(
                conn,
                candidate,
                "aspera_member",
                local_path,
                Path(package_path).name,
                source_url=url,
                package_path=package_path,
            )
            parse_harvested_file(conn, file_id, max_rows)
        except Exception as exc:  # noqa: BLE001
            insert_harvested_file(
                conn,
                candidate,
                "aspera_member",
                None,
                Path(package_path).name,
                source_url=url,
                package_path=package_path,
                status="error",
                error=str(exc),
            )


def write_meta(conn: sqlite3.Connection, key: str, value: Any) -> None:
    conn.execute(
        "INSERT OR REPLACE INTO harvest_meta (key, value) VALUES (?, ?)",
        (key, json_dumps(value)),
    )
    conn.commit()


def snapshot_meta(conn: sqlite3.Connection) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    try:
        rows = conn.execute("SELECT key, value FROM snapshot_meta")
    except sqlite3.Error:
        return meta
    for key, value in rows:
        try:
            meta[key] = json.loads(value)
        except (TypeError, json.JSONDecodeError):
            meta[key] = value
    return meta


def public_harvest_args(args: argparse.Namespace) -> dict[str, Any]:
    payload = dict(vars(args))
    path_keys = {
        "ascli",
        "files_dir",
        "out_db",
        "quality_flags_tsv",
        "source_db",
        "sums_inventory_tsv",
    }
    for key in path_keys:
        value = payload.get(key)
        if value:
            payload[key] = Path(str(value)).name
    return payload


def import_sums_inventory(conn: sqlite3.Connection, tsv_path: Path) -> None:
    if not tsv_path or not tsv_path.is_file():
        return
    conn.execute("DELETE FROM aspera_root_sums_inventory")
    with tsv_path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream, delimiter="\t")
        records = []
        dataset_type_by_key = {
            (row["short_title"], row["download_id"]): row["dataset_type"]
            for row in conn.execute(
                "SELECT short_title, download_id, dataset_type FROM candidate_downloads"
            )
        }
        for row in reader:
            records.append(
                (
                    row.get("short_title", ""),
                    dataset_type_by_key.get((row.get("short_title", ""), row.get("download_id", "")), ""),
                    row.get("download_id", ""),
                    row.get("download_title", ""),
                    row.get("sums_package_path", ""),
                    row.get("local_sums_path", ""),
                    int(row.get("line_number") or 0),
                    row.get("checksum", ""),
                    row.get("algorithm", ""),
                    clean_package_path(row.get("package_path", "")),
                    row.get("file_name", ""),
                    row.get("file_ext", "") or extension_for_name(row.get("package_path", "")),
                    row.get("raw_line", ""),
                )
            )
            if len(records) >= 10000:
                conn.executemany(
                    """
                    INSERT INTO aspera_root_sums_inventory
                    (short_title, dataset_type, download_id, download_title, sums_package_path,
                     local_sums_path, line_number, checksum, algorithm, package_path, file_name,
                     file_ext, raw_line)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    records,
                )
                records.clear()
        if records:
            conn.executemany(
                """
                INSERT INTO aspera_root_sums_inventory
                (short_title, dataset_type, download_id, download_title, sums_package_path,
                 local_sums_path, line_number, checksum, algorithm, package_path, file_name,
                 file_ext, raw_line)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                records,
            )
    conn.commit()


def import_quality_flags(conn: sqlite3.Connection, tsv_path: Path) -> None:
    if not tsv_path or not tsv_path.is_file():
        return
    with tsv_path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream, delimiter="\t")
        for row in reader:
            conn.execute(
                """
                INSERT INTO metadata_quality_flags
                (short_title, dataset_type, affected_table, affected_row_id, source_file_name,
                 sheet_name, source_row_number, excel_row, nifti_file, related_download_id,
                 issue_type, severity, status, evidence)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("short_title", ""),
                    "Collection",
                    "normalized_series_rows",
                    int(row.get("normalized_row_id") or 0),
                    row.get("source_workbook", ""),
                    row.get("sheet", ""),
                    int(row.get("source_row_number") or 0),
                    int(row.get("excel_row") or 0),
                    row.get("file_name", ""),
                    "",
                    "confirmed_missing_package_file",
                    "exclude_from_file_inventory",
                    "confirmed",
                    row.get("confirmation", ""),
                ),
            )
    conn.commit()


def first_nonempty(existing: str, candidate: str) -> str:
    return existing or candidate or ""


def build_nifti_file_series(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM nifti_file_series")
    candidate_meta = {
        (row["short_title"], row["download_id"]): dict(row)
        for row in conn.execute(
            """
            SELECT short_title, dataset_type, download_id, download_title
            FROM candidate_downloads
            """
        )
    }
    excluded = {
        (row["short_title"], clean_package_path(row["nifti_file"]))
        for row in conn.execute(
            """
            SELECT short_title, nifti_file
            FROM metadata_quality_flags
            WHERE status = 'confirmed'
              AND severity = 'exclude_from_file_inventory'
              AND NULLIF(nifti_file, '') IS NOT NULL
            """
        )
    }

    records: dict[tuple[str, str, str], dict[str, Any]] = {}

    def ensure(short_title: str, dataset_type: str, nifti_file: str) -> dict[str, Any]:
        cleaned = clean_package_path(nifti_file)
        key = (short_title, dataset_type, cleaned)
        record = records.get(key)
        if record is None:
            record = {
                "short_title": short_title,
                "dataset_type": dataset_type,
                "nifti_file": cleaned,
                "file_name": Path(cleaned).name,
                "file_ext": extension_for_name(cleaned),
                "package_path": cleaned,
                "download_ids": set(),
                "download_titles": set(),
                "inventory_sources": set(),
                "metadata_sources": set(),
                "source_row_count": 0,
                "checksum": "",
                "checksum_algorithm": "",
                "bytes": "",
                "quality_flag_json": "",
                "normalized": {name: "" for name in NORMALIZED_COLUMNS},
            }
            record["normalized"]["nifti_file"] = cleaned
            records[key] = record
        return record

    def apply_path_patient_inference(record: dict[str, Any], short_title: str, path: str) -> None:
        if record["normalized"].get("PatientID"):
            return
        patient_id, method = infer_patient_id_from_path(short_title, path)
        if not patient_id:
            return
        record["normalized"]["PatientID"] = patient_id
        record["metadata_sources"].add(f"path_patient_id:{method}")

    for row in conn.execute("SELECT * FROM normalized_series_rows"):
        nifti_file = clean_package_path(row["nifti_file"] or "")
        if not is_plausible_nifti_file(nifti_file):
            continue
        if (row["short_title"], nifti_file) in excluded:
            continue
        record = ensure(row["short_title"], row["dataset_type"], nifti_file)
        apply_path_patient_inference(record, row["short_title"], nifti_file)
        record["source_row_count"] += 1
        if row["download_id"]:
            record["download_ids"].add(row["download_id"])
        if row["download_title"]:
            record["download_titles"].add(row["download_title"])
        source = "normalized_series_rows"
        if row["source_file_name"]:
            source = f"{source}:{row['source_file_name']}:{row['sheet_name']}"
        record["metadata_sources"].add(source)
        for name in NORMALIZED_COLUMNS:
            record["normalized"][name] = first_nonempty(
                record["normalized"][name],
                row[name] or "",
            )

    for row in conn.execute("SELECT * FROM package_files"):
        package_path = clean_package_path(row["package_path"] or "")
        if not is_plausible_nifti_file(package_path):
            continue
        meta = candidate_meta.get((row["short_title"], row["download_id"]), {})
        dataset_type = meta.get("dataset_type", "")
        record = ensure(row["short_title"], dataset_type, package_path)
        apply_path_patient_inference(record, row["short_title"], package_path)
        record["source_row_count"] += 1
        record["download_ids"].add(row["download_id"] or "")
        record["download_titles"].add(row["download_title"] or "")
        record["inventory_sources"].add("package_files")
        record["bytes"] = first_nonempty(record["bytes"], row["bytes"] or "")

    for row in conn.execute("SELECT * FROM aspera_root_sums_inventory"):
        package_path = clean_package_path(row["package_path"] or "")
        if not is_plausible_nifti_file(package_path):
            continue
        dataset_type = row["dataset_type"] or candidate_meta.get(
            (row["short_title"], row["download_id"]), {}
        ).get("dataset_type", "")
        record = ensure(row["short_title"], dataset_type, package_path)
        apply_path_patient_inference(record, row["short_title"], package_path)
        record["source_row_count"] += 1
        record["download_ids"].add(row["download_id"] or "")
        record["download_titles"].add(row["download_title"] or "")
        record["inventory_sources"].add("aspera_root_sums_inventory")
        record["checksum"] = first_nonempty(record["checksum"], row["checksum"] or "")
        record["checksum_algorithm"] = first_nonempty(
            record["checksum_algorithm"], row["algorithm"] or ""
        )

    for row in conn.execute(
        """
        SELECT *
        FROM metadata_quality_flags
        WHERE status = 'confirmed'
          AND NULLIF(nifti_file, '') IS NOT NULL
        """
    ):
        key_candidates = [
            key for key in records if key[0] == row["short_title"] and key[2] == clean_package_path(row["nifti_file"])
        ]
        for key in key_candidates:
            records[key]["quality_flag_json"] = json_dumps(dict(row))

    normalized_placeholders = ",".join("?" for _ in NORMALIZED_COLUMNS)
    conn.executemany(
        f"""
        INSERT INTO nifti_file_series
        (short_title, dataset_type, download_ids, download_titles, file_name, file_ext,
         package_path, inventory_sources, metadata_sources, source_row_count, checksum,
         checksum_algorithm, bytes, quality_flag_json,
         {", ".join(f'"{name}"' for name in NORMALIZED_COLUMNS)})
        VALUES ({",".join("?" for _ in range(14))}, {normalized_placeholders})
        """,
        [
            (
                record["short_title"],
                record["dataset_type"],
                json_dumps(sorted(value for value in record["download_ids"] if value)),
                json_dumps(sorted(value for value in record["download_titles"] if value)),
                record["file_name"],
                record["file_ext"],
                record["package_path"],
                json_dumps(sorted(record["inventory_sources"])),
                json_dumps(sorted(record["metadata_sources"])),
                record["source_row_count"],
                record["checksum"],
                record["checksum_algorithm"],
                record["bytes"],
                record["quality_flag_json"],
                *[record["normalized"][name] for name in NORMALIZED_COLUMNS],
            )
            for record in records.values()
        ],
    )
    conn.commit()


def stable_hash_id(prefix: str, *values: Any) -> str:
    digest = hashlib.sha1(
        "\x1f".join(str(value or "") for value in values).encode("utf-8")
    ).hexdigest()[:20]
    return f"{prefix}_{digest}"


def parse_json_list(value: str) -> list[str]:
    try:
        parsed = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    return [str(item) for item in parsed if str(item)]


def strip_known_file_suffix(name: str) -> str:
    lower = name.lower()
    for suffix in (
        ".nii.gz",
        ".nii",
        ".eddy_rotated_bvecs",
        ".bvec",
        ".bval",
        ".json",
        ".txt",
        ".csv",
        ".tsv",
        ".xlsx",
        ".zip",
        ".sums",
    ):
        if lower.endswith(suffix):
            return name[: -len(suffix)]
    return Path(name).stem if "." in Path(name).name else name


def canonical_file_group_id(short_title: str, package_path: str) -> str:
    parent = str(Path(package_path).parent)
    if parent == ".":
        parent = ""
    stem = strip_known_file_suffix(Path(package_path).name)
    return stable_hash_id("fg", short_title, parent, stem)


def package_listing_row_is_directory(row: sqlite3.Row) -> bool:
    try:
        raw = json.loads(row["row_json"] or "{}")
    except json.JSONDecodeError:
        raw = {}
    return package_row_is_directory(raw)


def package_file_role(file_ext: str, file_name: str, package_path: str) -> str:
    lower = f"{file_name} {package_path}".lower()
    if file_ext in NIFTI_EXTENSIONS:
        return "image"
    if file_ext in {".bvec", ".bval", ".json", ".eddy_rotated_bvecs"}:
        return "sidecar"
    if file_ext in SPREADSHEET_EXTENSIONS or file_ext in {".txt", ".sums"}:
        return "package_metadata"
    if text_has_metadata_hint(lower):
        return "package_metadata"
    return "unknown"


def download_context(conn: sqlite3.Connection) -> dict[tuple[str, str], dict[str, Any]]:
    context: dict[tuple[str, str], dict[str, Any]] = {}
    for table in ("candidate_downloads", "nifti_downloads"):
        for row in conn.execute(
            f"""
            SELECT short_title, download_id, download_title, data_types,
                   file_types, download_types, license_label
            FROM {table}
            """
        ):
            key = (row["short_title"] or "", row["download_id"] or "")
            existing = context.setdefault(
                key,
                {
                    "download_title": "",
                    "data_types": set(),
                    "file_types": set(),
                    "download_types": set(),
                    "license_label": "",
                },
            )
            existing["download_title"] = existing["download_title"] or row["download_title"] or ""
            existing["license_label"] = existing["license_label"] or row["license_label"] or ""
            existing["data_types"].update(parse_json_list(row["data_types"] or "[]"))
            existing["file_types"].update(parse_json_list(row["file_types"] or "[]"))
            existing["download_types"].update(parse_json_list(row["download_types"] or "[]"))
    return context


def context_for_download_ids(
    context: dict[tuple[str, str], dict[str, Any]], short_title: str, download_ids_json: str
) -> dict[str, Any]:
    combined = {
        "download_title": "",
        "data_types": set(),
        "file_types": set(),
        "download_types": set(),
        "license_label": "",
    }
    for download_id in parse_json_list(download_ids_json):
        item = context.get((short_title, download_id))
        if not item:
            continue
        combined["download_title"] = combined["download_title"] or item["download_title"]
        combined["license_label"] = combined["license_label"] or item["license_label"]
        combined["data_types"].update(item["data_types"])
        combined["file_types"].update(item["file_types"])
        combined["download_types"].update(item["download_types"])
    return combined


def infer_modality(row: sqlite3.Row, context: dict[str, Any]) -> str:
    if row["Modality"]:
        return row["Modality"]
    data_types = {str(item).upper() for item in context["data_types"]}
    inferred: list[str] = []
    if "MR" in data_types:
        inferred.append("MR")
    if "CT" in data_types:
        inferred.append("CT")
    if "PT" in data_types or "PET" in data_types:
        inferred.append("PT")
    inferred = list(dict.fromkeys(inferred))
    if len(inferred) == 1:
        return inferred[0]
    return ""


def infer_derived_object_type(row: sqlite3.Row, context: dict[str, Any]) -> tuple[str, str, str]:
    text = " ".join(
        str(value or "")
        for value in (
            row["file_name"],
            row["SeriesDescription"],
            row["SegmentationType"],
            row["AlgorithmType"],
            row["AlgorithmName"],
        )
    ).lower()
    data_types = {str(item).lower() for item in context["data_types"]}
    if row["SegmentationType"] or "segmentation" in data_types:
        if re.search(r"(^|[_\-.])seg($|[_\-.])|segmentation|segment", text):
            return "segmentation", "segmentation_file", "segmentation metadata or filename"
        if "mask" in text:
            return "segmentation", "binary_mask", "segmentation data type and mask filename"
    if "labelmap" in text or "label_map" in text or re.search(r"(^|[_\-.])label", text):
        return "segmentation", "labelmap", "label filename"
    if "mask" in text:
        return "segmentation", "binary_mask", "mask filename"
    if re.search(r"(^|[_\-.])seg($|[_\-.])|segmentation|segment", text):
        return "segmentation", "segmentation_file", "segmentation filename"
    if "roi" in text:
        return "annotation", "roi", "roi filename"
    if "annotation" in text:
        return "annotation", "annotation", "annotation filename"
    if data_types and data_types.issubset({"segmentation", "image annotations", "annotation"}):
        return "segmentation", "unknown", "download data type"
    return "", "", ""


def infer_study_id(row: sqlite3.Row) -> tuple[str, str]:
    if row["StudyInstanceUID"]:
        return row["StudyInstanceUID"], "source_metadata"
    parent = clean_package_path(str(Path(row["package_path"] or "").parent))
    if parent:
        return stable_hash_id("study", row["short_title"], parent), "synthetic_from_parent_path"
    return stable_hash_id("study", row["short_title"], row["file_name"]), "synthetic_from_file_name"


def infer_series_id(row: sqlite3.Row, radiology_id: str) -> tuple[str, str]:
    if row["SeriesInstanceUID"]:
        return row["SeriesInstanceUID"], "source_metadata"
    return radiology_id, "synthetic_from_file_path"


def path_parent_text(package_path: str) -> str:
    parent = str(Path(package_path or "").parent)
    return "" if parent == "." else clean_package_path(parent)


def lower_file_stem(file_name: str) -> str:
    return strip_known_file_suffix(file_name or "").lower()


def reference_confidence_rank(confidence: str) -> int:
    return {"explicit": 4, "high": 3, "medium": 2, "low": 1}.get(confidence, 0)


def source_startswith(source: dict[str, Any], prefix: str) -> bool:
    source_stem = lower_file_stem(source["file_name"])
    prefix = prefix.lower()
    return source_stem == prefix or source_stem.startswith(f"{prefix}_") or source_stem.startswith(
        f"{prefix}-"
    )


def add_reference_candidates(
    candidates: dict[str, dict[str, Any]],
    derived: sqlite3.Row,
    sources: Iterable[dict[str, Any]],
    inference_method: str,
    confidence: str,
    evidence: dict[str, Any],
    reference_role: str = "source_image",
) -> None:
    for source in sources:
        key = source["radiology_id"]
        record = {
            "derived_object_id": derived["derived_object_id"],
            "derived_non_dicom_file_id": derived["non_dicom_file_id"],
            "derived_radiology_id": derived["radiology_id"],
            "referenced_non_dicom_file_id": source["non_dicom_file_id"],
            "referenced_radiology_id": source["radiology_id"],
            "referenced_series_id": source["series_id"],
            "referenced_file_name": source["file_name"],
            "referenced_package_path": source["package_path"],
            "reference_role": reference_role,
            "inference_method": inference_method,
            "confidence": confidence,
            "evidence_json": json_dumps(evidence),
        }
        existing = candidates.get(key)
        if existing is None or reference_confidence_rank(confidence) > reference_confidence_rank(
            existing["confidence"]
        ):
            candidates[key] = record


def exact_stem_matches(
    sources: Iterable[dict[str, Any]], target_stems: Iterable[str]
) -> list[dict[str, Any]]:
    targets = {target.lower() for target in target_stems if target}
    return [source for source in sources if lower_file_stem(source["file_name"]) in targets]


def prefix_matches(
    sources: Iterable[dict[str, Any]], prefix: str, max_sources: int = 80
) -> list[dict[str, Any]]:
    if not prefix:
        return []
    matches = [source for source in sources if source_startswith(source, prefix)]
    return matches if len(matches) <= max_sources else []


def infer_reference_candidates(
    derived: sqlite3.Row,
    sources_by_dataset: dict[str, list[dict[str, Any]]],
    sources_by_parent: dict[tuple[str, str], list[dict[str, Any]]],
    sources_by_series_id: dict[tuple[str, str], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    candidates: dict[str, dict[str, Any]] = {}
    short_title = derived["short_title"] or ""
    stem = strip_known_file_suffix(derived["file_name"] or "")
    stem_lower = stem.lower()
    package_path = clean_package_path(derived["package_path"] or "")
    package_path_lower = package_path.lower()
    parent = path_parent_text(package_path)
    same_parent = sources_by_parent.get((short_title, parent), [])
    dataset_sources = sources_by_dataset.get(short_title, [])

    referenced_series_id = derived["referenced_series_id"] or ""
    if referenced_series_id:
        add_reference_candidates(
            candidates,
            derived,
            sources_by_series_id.get((short_title, referenced_series_id), []),
            "explicit_referenced_series_id",
            "explicit",
            {"referenced_series_id": referenced_series_id},
        )

    label_match = re.fullmatch(r"labels[-_](.+)", stem_lower)
    if label_match:
        suffix = label_match.group(1)
        matches = exact_stem_matches(same_parent, [f"volume-{suffix}", f"volume_{suffix}"])
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "label_volume_filename_pair",
            "high",
            {"derived_stem": stem, "target_stems": [f"volume-{suffix}", f"volume_{suffix}"]},
        )

    if stem_lower == "lung_combined_mask" and "/tseg-lung-slow/" in package_path_lower:
        series_stem = Path(parent).name
        study_root = clean_package_path(package_path_lower.split("/tseg-lung-slow/", 1)[0])
        matches = [
            source
            for source in dataset_sources
            if lower_file_stem(source["file_name"]) == series_stem.lower()
            and clean_package_path(source["package_path"]).lower().startswith(study_root)
        ]
        original = [source for source in matches if "/resampled/" not in source["package_path"].lower()]
        resampled = [source for source in matches if "/resampled/" in source["package_path"].lower()]
        add_reference_candidates(
            candidates,
            derived,
            original,
            "nlst_lung_mask_parent_series",
            "high",
            {"series_stem_from_parent": series_stem, "study_root": study_root},
        )
        add_reference_candidates(
            candidates,
            derived,
            resampled,
            "nlst_lung_mask_parent_series_resampled",
            "medium",
            {"series_stem_from_parent": series_stem, "study_root": study_root},
        )

    seg_target_match = re.fullmatch(r"(.+?)_seg[_-]?(.+)", stem, flags=re.IGNORECASE)
    if seg_target_match:
        base, target = seg_target_match.groups()
        matches = exact_stem_matches(same_parent, [f"{base}_{target}", f"{base}-{target}", target])
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "segmentation_names_target_modality",
            "high",
            {"derived_stem": stem, "target_modality": target},
        )

    trailing_modality_match = re.fullmatch(r"(.+?)_([A-Za-z0-9]+)_seg", stem)
    if trailing_modality_match:
        base, target = trailing_modality_match.groups()
        matches = exact_stem_matches(same_parent, [f"{base}_{target}"])
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "trailing_modality_segmentation_pair",
            "high",
            {"derived_stem": stem, "target_stem": f"{base}_{target}"},
        )

    mask_match = re.match(r"(.+?)_mask(?:[_-].*)?$", stem, flags=re.IGNORECASE)
    if mask_match:
        prefix = mask_match.group(1)
        matches = prefix_matches(same_parent, prefix)
        confidence = "high" if len(matches) == 1 else "medium"
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "mask_prefix_same_folder",
            confidence,
            {"derived_stem": stem, "source_prefix": prefix, "same_parent_source_count": len(matches)},
        )

    prefix_patterns = [
        (r"(.+?)_(?:brats_tumor|enhancing_cellular_tumor|non_enhancing_cellular_tumor|total_cellular_tumor)_seg$", "tumor_segmentation_prefix"),
        (r"(.+?)_(?:core|whole)_seg$", "core_whole_segmentation_prefix"),
        (r"(.+?)_(?:brain(?:_parenchyma)?|tumor)_segmentation$", "anatomic_segmentation_prefix"),
        (r"(.+?)_(?:lesions|spine)_segmentation$", "lesion_spine_segmentation_prefix"),
        (r"(.+?)_tumormask$", "tumor_mask_prefix"),
        (r"(.+?)[_-](?:seg|segmentation|segmentations)$", "segmentation_suffix_prefix"),
    ]
    for pattern, method in prefix_patterns:
        match = re.fullmatch(pattern, stem, flags=re.IGNORECASE)
        if not match:
            continue
        prefix = match.group(1)
        matches = prefix_matches(same_parent, prefix)
        confidence = "high" if len(matches) == 1 else "medium"
        add_reference_candidates(
            candidates,
            derived,
            matches,
            method,
            confidence,
            {"derived_stem": stem, "source_prefix": prefix, "same_parent_source_count": len(matches)},
        )

    ivy_match = re.fullmatch(r"(.+?)_[^_]+_labels", stem, flags=re.IGNORECASE)
    if "ivygap-radiomics" in short_title.lower() and ivy_match:
        prefix = ivy_match.group(1)
        atlas = ""
        if "annotations_mni" in package_path_lower:
            atlas = "images_mni"
        elif "annotations_sri" in package_path_lower:
            atlas = "images_sri"
        matches = [
            source
            for source in dataset_sources
            if source_startswith(source, prefix)
            and (not atlas or atlas in source["package_path"].lower())
        ]
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "ivygap_label_prefix_and_atlas",
            "medium",
            {"derived_stem": stem, "source_prefix": prefix, "atlas_hint": atlas},
        )

    brats_tcga_match = re.fullmatch(
        r"(TCGA-[A-Z0-9]{2}-[A-Z0-9]{4}_\d{4}\.\d{2}\.\d{2})_GlistrBoost(?:_ManuallyCorrected)?",
        stem,
        flags=re.IGNORECASE,
    )
    if short_title in {"BraTS-TCGA-GBM", "BraTS-TCGA-LGG"} and brats_tcga_match:
        prefix = brats_tcga_match.group(1)
        matches = [source for source in same_parent if source_startswith(source, prefix)]
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "brats_tcga_glistrboost_same_subject_date",
            "high",
            {"derived_stem": stem, "source_prefix": prefix, "same_parent_source_count": len(matches)},
        )

    brats_peds_match = re.fullmatch(
        r"(BraTS-PED-\d{5}-\d{3})-seg",
        stem,
        flags=re.IGNORECASE,
    )
    if short_title == "BraTS-PEDs" and brats_peds_match:
        prefix = brats_peds_match.group(1)
        matches = [source for source in same_parent if source_startswith(source, prefix)]
        add_reference_candidates(
            candidates,
            derived,
            matches,
            "brats_peds_seg_same_subject_timepoint",
            "high",
            {"derived_stem": stem, "source_prefix": prefix, "same_parent_source_count": len(matches)},
        )

    if not candidates and 0 < len(same_parent) <= 50:
        add_reference_candidates(
            candidates,
            derived,
            same_parent,
            "same_folder_source_images",
            "low",
            {"parent_path": parent, "same_parent_source_count": len(same_parent)},
        )

    return sorted(
        candidates.values(),
        key=lambda item: (
            -reference_confidence_rank(item["confidence"]),
            item["referenced_package_path"],
            item["referenced_file_name"],
        ),
    )


def link_derived_object_references(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM derived_object_references")
    source_rows = [
        dict(row)
        for row in conn.execute(
            """
            SELECT radiology_id, non_dicom_file_id, short_title, file_name,
                   package_path, series_id
            FROM radiology_series
            WHERE is_derived_object = 0
            """
        )
    ]
    sources_by_dataset: dict[str, list[dict[str, Any]]] = {}
    sources_by_parent: dict[tuple[str, str], list[dict[str, Any]]] = {}
    sources_by_series_id: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for source in source_rows:
        short_title = source["short_title"] or ""
        sources_by_dataset.setdefault(short_title, []).append(source)
        sources_by_parent.setdefault((short_title, path_parent_text(source["package_path"])), []).append(
            source
        )
        if source["series_id"]:
            sources_by_series_id.setdefault((short_title, source["series_id"]), []).append(source)

    reference_records = []
    best_reference_updates = []
    for derived in conn.execute("SELECT * FROM derived_objects"):
        candidates = infer_reference_candidates(
            derived, sources_by_dataset, sources_by_parent, sources_by_series_id
        )
        for candidate in candidates:
            reference_records.append(
                (
                    stable_hash_id(
                        "dor",
                        candidate["derived_object_id"],
                        candidate["referenced_non_dicom_file_id"],
                        candidate["inference_method"],
                    ),
                    candidate["derived_object_id"],
                    candidate["derived_non_dicom_file_id"],
                    candidate["derived_radiology_id"],
                    candidate["referenced_non_dicom_file_id"],
                    candidate["referenced_radiology_id"],
                    candidate["referenced_series_id"],
                    candidate["referenced_file_name"],
                    candidate["referenced_package_path"],
                    candidate["reference_role"],
                    candidate["inference_method"],
                    candidate["confidence"],
                    candidate["evidence_json"],
                )
            )
        if candidates:
            best_rank = reference_confidence_rank(candidates[0]["confidence"])
            best_candidates = [
                candidate
                for candidate in candidates
                if reference_confidence_rank(candidate["confidence"]) == best_rank
            ]
            if len(best_candidates) == 1:
                best = best_candidates[0]
                best_reference_updates.append(
                    (
                        best["referenced_radiology_id"],
                        best["referenced_series_id"],
                        derived["derived_object_id"],
                    )
                )

    conn.executemany(
        """
        INSERT OR REPLACE INTO derived_object_references
        (derived_object_reference_id, derived_object_id, derived_non_dicom_file_id,
         derived_radiology_id, referenced_non_dicom_file_id, referenced_radiology_id,
         referenced_series_id, referenced_file_name, referenced_package_path, reference_role,
         inference_method, confidence, evidence_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        reference_records,
    )
    conn.executemany(
        """
        UPDATE derived_objects
        SET referenced_radiology_id = ?, referenced_series_id = ?
        WHERE derived_object_id = ?
        """,
        best_reference_updates,
    )


def insert_non_dicom_file_records(conn: sqlite3.Connection) -> None:
    records = []
    context_by_download = download_context(conn)
    for row in conn.execute("SELECT * FROM nifti_file_series"):
        package_path = clean_package_path(row["package_path"] or row["nifti_file"] or "")
        file_group_id = canonical_file_group_id(row["short_title"], package_path)
        context = context_for_download_ids(context_by_download, row["short_title"], row["download_ids"])
        derived_type, _representation, _basis = infer_derived_object_type(row, context)
        records.append(
            (
                stable_hash_id("file", row["short_title"], row["dataset_type"], package_path),
                row["short_title"],
                row["dataset_type"],
                row["download_ids"],
                row["download_titles"],
                row["file_name"],
                row["file_ext"],
                package_path,
                file_group_id,
                "image",
                row["inventory_sources"],
                row["metadata_sources"],
                "nifti_file_series",
                row["nifti_file_series_id"],
                row["bytes"],
                row["checksum"],
                row["checksum_algorithm"],
                1,
                0,
                0,
                1 if derived_type else 0,
                row["quality_flag_json"],
            )
        )

    candidate_meta = {
        (row["short_title"], row["download_id"]): dict(row)
        for row in conn.execute("SELECT * FROM candidate_downloads")
    }
    for row in conn.execute("SELECT * FROM package_files WHERE file_ext NOT IN ('.nii', '.nii.gz')"):
        if package_listing_row_is_directory(row):
            continue
        package_path = clean_package_path(row["package_path"] or "")
        if not package_path:
            continue
        meta = candidate_meta.get((row["short_title"], row["download_id"]), {})
        dataset_type = meta.get("dataset_type", "")
        file_role = package_file_role(row["file_ext"] or "", row["file_name"] or "", package_path)
        is_sidecar = 1 if file_role == "sidecar" else 0
        records.append(
            (
                stable_hash_id("file", row["short_title"], dataset_type, package_path),
                row["short_title"],
                dataset_type,
                json_dumps([row["download_id"]]) if row["download_id"] else "[]",
                json_dumps([row["download_title"]]) if row["download_title"] else "[]",
                row["file_name"],
                row["file_ext"],
                package_path,
                canonical_file_group_id(row["short_title"], package_path),
                file_role,
                json_dumps(["package_files"]),
                "[]",
                "package_files",
                row["package_file_id"],
                row["bytes"],
                "",
                "",
                0,
                is_sidecar,
                1 if file_role == "package_metadata" else 0,
                0,
                "",
            )
        )
    conn.executemany(
        """
        INSERT OR REPLACE INTO non_dicom_files
        (non_dicom_file_id, short_title, dataset_type, download_ids, download_titles,
         file_name, file_ext, package_path, file_group_id, file_role,
         inventory_sources, metadata_sources, source_table, source_row_id, bytes,
         checksum, checksum_algorithm, is_nifti, is_sidecar, is_package_metadata,
         is_derived_candidate, quality_flag_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        records,
    )


def build_canonical_non_dicom_layer(conn: sqlite3.Connection) -> None:
    for table in (
        "annotation_groups",
        "derived_object_references",
        "derived_objects",
        "radiology_contrast",
        "radiology_pet",
        "radiology_ct",
        "radiology_mr",
        "radiology_series",
        "non_dicom_files",
    ):
        conn.execute(f"DELETE FROM {table}")
    insert_non_dicom_file_records(conn)

    context = download_context(conn)
    file_id_by_source = {
        (row["source_table"], row["source_row_id"]): row["non_dicom_file_id"]
        for row in conn.execute("SELECT source_table, source_row_id, non_dicom_file_id FROM non_dicom_files")
    }
    file_group_by_id = {
        row["non_dicom_file_id"]: row["file_group_id"]
        for row in conn.execute("SELECT non_dicom_file_id, file_group_id FROM non_dicom_files")
    }

    radiology_records = []
    mr_records = []
    ct_records = []
    pet_records = []
    contrast_records = []
    derived_records = []

    for row in conn.execute("SELECT * FROM nifti_file_series"):
        non_dicom_file_id = file_id_by_source[("nifti_file_series", row["nifti_file_series_id"])]
        file_group_id = file_group_by_id[non_dicom_file_id]
        radiology_id = stable_hash_id("rad", row["short_title"], row["dataset_type"], row["package_path"])
        row_context = context_for_download_ids(context, row["short_title"], row["download_ids"])
        modality = infer_modality(row, row_context)
        derived_type, segmentation_representation, derivation_basis = infer_derived_object_type(
            row, row_context
        )
        study_id, study_id_source = infer_study_id(row)
        series_id, series_id_source = infer_series_id(row, radiology_id)
        object_type = derived_type or "NIfTI image"
        radiology_records.append(
            (
                radiology_id,
                non_dicom_file_id,
                file_group_id,
                row["short_title"],
                row["dataset_type"],
                row["download_ids"],
                row["file_name"],
                row["package_path"],
                row["PatientID"],
                "",
                study_id,
                study_id_source,
                series_id,
                series_id_source,
                row["source_DOI"],
                modality,
                row["BodyPartExamined"],
                row["StudyDate"],
                row["SeriesDate"],
                row["StudyDescription"],
                row["SeriesDescription"],
                row["SeriesNumber"],
                row["Manufacturer"],
                row["ManufacturerModelName"],
                "",
                "",
                object_type,
                row["Rows"],
                row["Columns"],
                "",
                row["NumberOfTemporalPositions"],
                row["PixelSpacing_row_mm"],
                row["PixelSpacing_col_mm"],
                row["SliceThickness"],
                "",
                "",
                "",
                1 if derived_type else 0,
                row["quality_flag_json"],
            )
        )
        if modality == "MR" or any(row[name] for name in (
            "MagneticFieldStrength",
            "ScanningSequence",
            "SequenceVariant",
            "MRAcquisitionType",
            "EchoTime",
            "RepetitionTime",
            "EchoTrainLength",
            "FlipAngle",
            "InversionTime",
            "ReceiveCoilName",
            "SequenceName",
            "DiffusionBValue",
        )):
            mr_records.append(
                (
                    radiology_id,
                    row["MagneticFieldStrength"],
                    row["ScanningSequence"],
                    row["SequenceVariant"],
                    row["MRAcquisitionType"],
                    row["EchoTime"],
                    row["RepetitionTime"],
                    row["EchoTrainLength"],
                    row["FlipAngle"],
                    "",
                    "",
                    "",
                    row["InversionTime"],
                    row["ReceiveCoilName"],
                    row["SequenceName"],
                    row["DiffusionBValue"],
                )
            )
        if modality == "CT" or any(row[name] for name in (
            "KVP",
            "ScanOptions",
            "ConvolutionKernel",
            "GantryDetectorTilt",
            "XRayTubeCurrent_min",
            "XRayTubeCurrent_max",
            "FilterType",
            "Exposure_min",
            "Exposure_max",
            "ExposureTime_min",
            "ExposureTime_max",
            "DataCollectionDiameter",
            "ReconstructionDiameter",
            "SpiralPitchFactor",
        )):
            ct_records.append(
                (
                    radiology_id,
                    row["KVP"],
                    row["ScanOptions"],
                    row["ConvolutionKernel"],
                    row["GantryDetectorTilt"],
                    row["XRayTubeCurrent_min"],
                    row["XRayTubeCurrent_max"],
                    row["FilterType"],
                    row["Exposure_min"],
                    row["Exposure_max"],
                    row["ExposureTime_min"],
                    row["ExposureTime_max"],
                    row["DataCollectionDiameter"],
                    row["ReconstructionDiameter"],
                    row["SpiralPitchFactor"],
                )
            )
        if modality == "PT":
            pet_records.append((radiology_id, "", "", "", "", "", "", "", "", "", "", "", "", "", ""))
        if any(row[name] for name in (
            "ContrastBolusAgent",
            "ContrastBolusIngredient",
            "ContrastBolusRoute",
        )):
            contrast_records.append(
                (
                    radiology_id,
                    row["ContrastBolusAgent"],
                    row["ContrastBolusIngredient"],
                    row["ContrastBolusRoute"],
                )
            )
        if derived_type:
            derived_records.append(
                (
                    stable_hash_id("der", row["short_title"], row["dataset_type"], row["package_path"]),
                    non_dicom_file_id,
                    radiology_id,
                    row["short_title"],
                    row["dataset_type"],
                    row["file_name"],
                    row["package_path"],
                    row["file_ext"],
                    row["analysis_result_id"],
                    "",
                    row["segmented_SeriesInstanceUID"],
                    derived_type,
                    segmentation_representation,
                    row["SegmentationType"],
                    row["total_segments"],
                    row["AlgorithmType"],
                    row["AlgorithmName"],
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    derivation_basis,
                )
            )

    conn.executemany(
        """
        INSERT INTO radiology_series
        (radiology_id, non_dicom_file_id, file_group_id, short_title, dataset_type,
         download_ids, file_name, package_path, subject_id, procedure_id, study_id,
         study_id_source, series_id, series_id_source, source_doi, modality,
         body_part_examined, study_date, series_date, study_description, series_description,
         series_number, manufacturer, manufacturer_model_name, software_versions, image_type,
         object_type, rows, columns, number_of_slices, number_of_temporal_positions,
         pixel_spacing_row_mm, pixel_spacing_col_mm, slice_thickness_mm,
         spacing_between_slices_mm, orientation_or_affine, is_phantom, is_derived_object,
         quality_flag_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        radiology_records,
    )
    conn.executemany(
        """
        INSERT INTO radiology_mr
        (radiology_id, magnetic_field_strength_t, scanning_sequence, sequence_variant,
         mr_acquisition_type, echo_time_ms, repetition_time_ms, echo_train_length,
         flip_angle_deg, pixel_bandwidth_hz, imaging_frequency_mhz, imaged_nucleus,
         inversion_time_ms, receive_coil_name, sequence_name, diffusion_b_value_s_per_mm2)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        mr_records,
    )
    conn.executemany(
        """
        INSERT INTO radiology_ct
        (radiology_id, kvp, scan_options, convolution_kernel, gantry_detector_tilt_deg,
         xray_tube_current_min_ma, xray_tube_current_max_ma, filter_type,
         exposure_min_mas, exposure_max_mas, exposure_time_min_ms,
         exposure_time_max_ms, data_collection_diameter_mm,
         reconstruction_diameter_mm, spiral_pitch_factor)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        ct_records,
    )
    conn.executemany(
        """
        INSERT INTO radiology_pet
        (radiology_id, series_type, units, decay_correction, corrected_image,
         randoms_correction_method, reconstruction_method, actual_frame_duration_ms,
         scatter_correction_method, attenuation_correction_method, radionuclide_code_meaning,
         radionuclide_total_dose_bq, radiopharmaceutical_start_time, radiopharmaceutical,
         number_of_time_slices)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        pet_records,
    )
    conn.executemany(
        """
        INSERT INTO radiology_contrast
        (radiology_id, contrast_bolus_agent, contrast_bolus_ingredient, contrast_bolus_route)
        VALUES (?, ?, ?, ?)
        """,
        contrast_records,
    )
    conn.executemany(
        """
        INSERT INTO derived_objects
        (derived_object_id, non_dicom_file_id, radiology_id, short_title, dataset_type,
         file_name, package_path, file_ext, analysis_result_id, referenced_radiology_id,
         referenced_series_id, derived_object_type, segmentation_representation,
         segmentation_type, total_segments, algorithm_type, algorithm_name,
         segmented_property_category, segmented_property_type, anatomic_region, roi_names,
         roi_generation_algorithms, rt_roi_interpreted_types, annotation_coordinate_type,
         derivation_basis)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        derived_records,
    )
    link_derived_object_references(conn)
    conn.commit()


def append_json_list_value(value: str, item: str) -> str:
    values = parse_json_list(value)
    if item not in values:
        values.append(item)
    return json_dumps(sorted(values))


def resolve_harvested_file_path(file_meta: sqlite3.Row, files_dir: Path) -> Path | None:
    candidates: list[Path] = []
    local_path = str(file_meta["local_path"] or "")
    if local_path:
        candidates.append(Path(local_path))
    candidates.append(candidate_local_path(files_dir, dict(file_meta), file_meta["file_name"] or "download"))
    for path in candidates:
        if path.exists():
            return path
    glob_root = files_dir / safe_slug(file_meta["short_title"] or "dataset")
    if glob_root.exists():
        matches = sorted(glob_root.glob(f"*/{safe_slug(file_meta['file_name'] or '', 'download')}"))
        for path in matches:
            if path.exists():
                return path
    return None


def harvested_support_file(
    conn: sqlite3.Connection, short_title: str, file_name: str, files_dir: Path
) -> Path | None:
    for row in conn.execute(
        """
        SELECT *
        FROM harvested_files
        WHERE short_title = ?
          AND file_name = ?
          AND status = 'ok'
        ORDER BY file_id DESC
        """,
        (short_title, file_name),
    ):
        path = resolve_harvested_file_path(row, files_dir)
        if path:
            return path
    return None


def find_local_support_file(files_dir: Path, file_name: str) -> Path | None:
    for base in local_support_search_bases(files_dir):
        candidates = [base] if base.name == file_name else [base / file_name]
        candidates.extend(
            [
                base / "outputs" / "nifti_metadata_candidate" / file_name,
                base / "outputs" / "nifti_metadata" / file_name,
                base / "outputs" / "nifti_metadata_full" / file_name,
            ]
        )
        for candidate in candidates:
            if candidate.is_file():
                return candidate
    return None


def parse_tcia_manifest_series_uids(path: Path) -> list[str]:
    uid_re = re.compile(r"\b\d+(?:\.\d+)+\b")
    uids: list[str] = []
    seen: set[str] = set()
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith(("#", ";")):
            continue
        for match in uid_re.finditer(stripped):
            uid = match.group(0)
            if 10 <= len(uid) <= 64 and uid not in seen:
                seen.add(uid)
                uids.append(uid)
    return uids


def inventory_cached_direct_zip_nifti_members(conn: sqlite3.Connection, files_dir: Path) -> int:
    inserted = 0
    for row in conn.execute(
        """
        SELECT
            h.*,
            c.download_row_id AS candidate_download_row_id,
            c.dataset_type AS candidate_dataset_type,
            c.download_title AS candidate_download_title,
            c.download_url AS candidate_download_url
        FROM harvested_files h
        LEFT JOIN candidate_downloads c
          ON c.short_title = h.short_title
         AND c.download_id = h.download_id
        WHERE h.source_kind = 'direct_zip'
          AND h.status = 'ok'
        """
    ):
        zip_path = resolve_harvested_file_path(row, files_dir)
        if not zip_path:
            continue
        candidate = {
            "download_row_id": row["candidate_download_row_id"] or row["download_row_id"],
            "short_title": row["short_title"],
            "dataset_type": row["candidate_dataset_type"] or row["dataset_type"],
            "download_id": row["download_id"],
            "download_title": row["candidate_download_title"] or row["download_title"],
            "download_url": row["candidate_download_url"] or row["source_url"],
        }
        inserted += insert_direct_zip_nifti_package_files(conn, candidate, zip_path)
    return inserted


def parse_brats2021_mapping(path: Path) -> dict[str, str]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse BraTS2021_MappingToTCIA.xlsx") from exc
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    mapping: dict[str, str] = {}
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for index, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if index == 1:
                continue
            tcia_patient_id = stringify_cell(row[2] if len(row) > 2 else "").strip()
            challenge_id = stringify_cell(row[4] if len(row) > 4 else "").strip()
            if tcia_patient_id and challenge_id:
                if normalize_key(tcia_patient_id) in {"newnotpreviouslyintcia", "newnotintcia"}:
                    tcia_patient_id = challenge_id
                mapping[challenge_id] = tcia_patient_id
    finally:
        workbook.close()
    return mapping


def parse_csv_dicts(path: Path, delimiter: str | None = None) -> list[dict[str, str]]:
    _headers, rows = parse_csv_like(path, delimiter=delimiter)
    return rows


def update_nifti_series_fields(
    conn: sqlite3.Connection,
    nifti_file_series_id: int,
    updates: dict[str, str],
    metadata_source: str,
) -> None:
    row = conn.execute(
        """
        SELECT metadata_sources
        FROM nifti_file_series
        WHERE nifti_file_series_id = ?
        """,
        (nifti_file_series_id,),
    ).fetchone()
    if row is None:
        return
    updates = {key: value for key, value in updates.items() if value is not None}
    if not updates:
        return
    updates["metadata_sources"] = append_json_list_value(row["metadata_sources"], metadata_source)
    assignments = ", ".join(f'"{key}" = ?' for key in updates)
    conn.execute(
        f"UPDATE nifti_file_series SET {assignments} WHERE nifti_file_series_id = ?",
        (*updates.values(), nifti_file_series_id),
    )


def apply_rsna_brats2021_mapping(conn: sqlite3.Connection, files_dir: Path) -> int:
    path = harvested_support_file(
        conn,
        "RSNA-ASNR-MICCAI-BraTS-2021",
        "BraTS2021_MappingToTCIA.xlsx",
        files_dir,
    )
    if not path:
        return 0
    mapping = parse_brats2021_mapping(path)
    updates = 0
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, PatientID, file_name, package_path
        FROM nifti_file_series
        WHERE short_title = 'RSNA-ASNR-MICCAI-BraTS-2021'
        """
    ):
        match = re.search(r"(BraTS2021_\d{5})", row["package_path"] or row["file_name"] or "")
        if not match:
            continue
        patient_id = mapping.get(match.group(1), "")
        if not patient_id or row["PatientID"] == patient_id:
            continue
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {"PatientID": patient_id},
            "dataset_specific:BraTS2021_MappingToTCIA.xlsx:column_C_by_column_E",
        )
        updates += 1
    return updates


def apply_saros_mapping(conn: sqlite3.Connection, files_dir: Path) -> int:
    path = harvested_support_file(conn, "SAROS", "Segmentation-Info_09-29-2023.csv", files_dir)
    if not path:
        return 0
    rows = parse_csv_dicts(path)
    mapping = {
        (row.get("id") or "").strip(): row
        for row in rows
        if (row.get("id") or "").strip()
    }
    updates = 0
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name, package_path
        FROM nifti_file_series
        WHERE short_title = 'SAROS'
        """
    ):
        match = re.search(r"(case_\d{3})", row["package_path"] or row["file_name"] or "")
        if not match:
            continue
        source = mapping.get(match.group(1))
        if not source:
            continue
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {
                "PatientID": (source.get("tcia_case_id") or "").strip(),
                "StudyInstanceUID": (source.get("tcia_study_instance_uid") or "").strip(),
                "segmented_SeriesInstanceUID": (source.get("tcia_series_instance_uid") or "").strip(),
                "PatientSex": (source.get("gender") or "").strip(),
                "BodyPartExamined": (source.get("anatomic_region") or "").strip(),
            },
            "dataset_specific:Segmentation-Info_09-29-2023.csv:id_to_tcia_case_id",
        )
        updates += 1
    return updates


def apply_breastdcedl_ispy2_mapping(conn: sqlite3.Connection, files_dir: Path) -> int:
    path = harvested_support_file(
        conn,
        "BreastDCEDL_ISPY2",
        "BreastDCEDL_ISPY2_files_sizes.tsv",
        files_dir,
    )
    if not path:
        return 0
    rows = parse_csv_dicts(path, delimiter="\t")
    mapping = {
        (row.get("filename") or "").strip(): row
        for row in rows
        if (row.get("filename") or "").strip()
    }
    updates = 0
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name, PatientID
        FROM nifti_file_series
        WHERE short_title = 'BreastDCEDL_ISPY2'
        """
    ):
        source = mapping.get(row["file_name"] or "")
        if not source:
            continue
        patient_id = (source.get("pid") or "").strip()
        if not patient_id:
            continue
        field_updates = {"PatientID": patient_id}
        file_size = (source.get("file_size") or "").strip()
        if file_size:
            field_updates["bytes"] = file_size
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            field_updates,
            "dataset_specific:BreastDCEDL_ISPY2_files_sizes.tsv:filename_to_pid",
        )
        updates += 1
    return updates


def apply_ct_org_mapping(conn: sqlite3.Connection) -> int:
    updates = 0
    pattern = re.compile(r"^(?:volume|labels)-(\d+)\.nii(?:\.gz)?$", flags=re.IGNORECASE)
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name
        FROM nifti_file_series
        WHERE short_title = 'CT-ORG'
        """
    ):
        match = pattern.match(row["file_name"] or "")
        if not match:
            continue
        case_number = match.group(1)
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {"PatientID": f"CT-ORG-{case_number}", "Modality": "CT"},
            "dataset_specific:CT-ORG_readme_volume_labels_case_number",
        )
        updates += 1
    return updates


def idc_series_metadata_for_uids(uids: list[str]) -> dict[str, dict[str, str]]:
    if not uids:
        return {}
    try:
        from idc_index import IDCClient  # type: ignore
    except ImportError:
        return {}
    client = IDCClient()
    quoted_uids = ",".join("'" + uid.replace("'", "''") + "'" for uid in uids)
    query = f"""
        SELECT collection_id, PatientID, StudyInstanceUID, SeriesInstanceUID, Modality,
               BodyPartExamined, StudyDate, StudyDescription, SeriesDescription,
               instanceCount, license_short_name, source_DOI
        FROM index
        WHERE SeriesInstanceUID IN ({quoted_uids})
    """
    rows = client.sql_query(query).to_dict("records")
    by_patient: dict[str, dict[str, str]] = {}
    for row in rows:
        record = {key: "" if value is None else str(value) for key, value in row.items()}
        patient_id = record.get("PatientID", "")
        if patient_id:
            by_patient[patient_id] = record
    return by_patient


def apply_pancreas_ct_mapping(conn: sqlite3.Connection, files_dir: Path) -> int:
    manifest_path = find_local_support_file(files_dir, PANCREAS_CT_MANIFEST_FILE)
    if not manifest_path:
        return 0
    uids = parse_tcia_manifest_series_uids(manifest_path)
    try:
        source_by_patient = idc_series_metadata_for_uids(uids)
    except Exception as exc:  # noqa: BLE001 - IDC enrichment is optional at postprocess time.
        write_meta(conn, "pancreas_ct_idc_mapping_error", str(exc))
        return 0
    if not source_by_patient:
        return 0
    updates = 0
    pattern = re.compile(r"^label(\d{4})\.nii(?:\.gz)?$", flags=re.IGNORECASE)
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name
        FROM nifti_file_series
        WHERE short_title = 'Pancreas-CT'
        """
    ):
        match = pattern.match(row["file_name"] or "")
        if not match:
            continue
        patient_id = f"PANCREAS_{int(match.group(1)):04d}"
        source = source_by_patient.get(patient_id)
        if not source:
            continue
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {
                "collection_id": source.get("collection_id", ""),
                "PatientID": patient_id,
                "StudyInstanceUID": source.get("StudyInstanceUID", ""),
                "source_DOI": source.get("source_DOI", ""),
                "StudyDate": source.get("StudyDate", ""),
                "SeriesDate": source.get("StudyDate", ""),
                "StudyDescription": source.get("StudyDescription", ""),
                "BodyPartExamined": source.get("BodyPartExamined", "") or "PANCREAS",
                "Modality": source.get("Modality", "") or "CT",
                "SeriesDescription": "manual pancreas segmentation",
                "SegmentationType": "labelmap",
                "total_segments": "1",
                "AlgorithmType": "MANUAL",
                "AlgorithmName": "manual pancreas annotation",
                "segmented_SeriesInstanceUID": source.get("SeriesInstanceUID", ""),
                "license_short_name": source.get("license_short_name", ""),
            },
            "dataset_specific:Pancreas-CT_manifest_idc_patient_suffix",
        )
        updates += 1
    write_meta(
        conn,
        "pancreas_ct_manifest_idc_summary",
        {
            "manifest_file": PANCREAS_CT_MANIFEST_FILE,
            "manifest_series_instance_uids": len(uids),
            "idc_patient_rows": len(source_by_patient),
        },
    )
    return updates


def apply_brats_tcga_mapping(conn: sqlite3.Connection) -> int:
    updates = 0
    pattern = re.compile(
        r"^(TCGA-[A-Z0-9]{2}-[A-Z0-9]{4})_(\d{4})\.(\d{2})\.(\d{2})_(.+?)\.nii(?:\.gz)?$",
        flags=re.IGNORECASE,
    )
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, short_title, file_name
        FROM nifti_file_series
        WHERE short_title IN ('BraTS-TCGA-GBM', 'BraTS-TCGA-LGG')
        """
    ):
        match = pattern.match(row["file_name"] or "")
        if not match:
            continue
        patient_id, year, month, day, scan_type = match.groups()
        scan_type_lower = scan_type.lower()
        field_updates = {
            "PatientID": patient_id,
            "StudyDate": f"{year}-{month}-{day}",
            "SeriesDate": f"{year}-{month}-{day}",
            "SeriesDescription": scan_type,
            "BodyPartExamined": "BRAIN",
            "Modality": "MR",
        }
        if scan_type_lower.startswith("glistrboost"):
            field_updates.update(
                {
                    "SegmentationType": "labelmap",
                    "AlgorithmName": "GLISTRboost",
                    "AlgorithmType": "SEMIAUTOMATIC"
                    if "manuallycorrected" in normalize_key(scan_type)
                    else "AUTOMATIC",
                }
            )
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            field_updates,
            "dataset_specific:BraTS-TCGA_filename_patient_date_scan_type",
        )
        updates += 1
    return updates


def apply_brats_peds_mapping(conn: sqlite3.Connection) -> int:
    updates = 0
    pattern = re.compile(
        r"^(BraTS-PED-\d{5}-\d{3})-(t1n|t1c|t2w|t2f|seg)\.nii(?:\.gz)?$",
        flags=re.IGNORECASE,
    )
    sequence_labels = {
        "t1n": "native T1",
        "t1c": "contrast-enhanced T1",
        "t2w": "T2-weighted",
        "t2f": "T2-FLAIR",
        "seg": "expert-refined tumor segmentation",
    }
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name
        FROM nifti_file_series
        WHERE short_title = 'BraTS-PEDs'
        """
    ):
        match = pattern.match(row["file_name"] or "")
        if not match:
            continue
        patient_id, scan_type = match.groups()
        scan_type = scan_type.lower()
        field_updates = {
            "PatientID": patient_id,
            "BodyPartExamined": "BRAIN",
            "Modality": "MR",
            "SeriesDescription": sequence_labels[scan_type],
        }
        if scan_type == "seg":
            field_updates.update(
                {
                    "SegmentationType": "labelmap",
                    "AlgorithmType": "MANUAL",
                    "AlgorithmName": "expert-refined tumor segmentation",
                }
            )
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            field_updates,
            "dataset_specific:BraTS-PEDs_filename_subject_timepoint_scan_type",
        )
        updates += 1
    return updates


def apply_plethora_mapping(conn: sqlite3.Connection) -> int:
    updates = 0
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name, package_path
        FROM nifti_file_series
        WHERE short_title = 'PleThora'
        """
    ):
        match = re.search(r"(LUNG1-\d{3})", row["package_path"] or row["file_name"] or "")
        if not match:
            continue
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {"PatientID": match.group(1)},
            "dataset_specific:PleThora_zip_directory_patient_id",
        )
        updates += 1
    return updates


def apply_vs_mc_rc2_mapping(conn: sqlite3.Connection) -> int:
    updates = 0
    pattern = re.compile(
        r"^(VS_MC_RC2_\d{3})_(\d{4}-\d{2}-\d{2})_([A-Za-z0-9]+(?:_seg)?)\.nii(?:\.gz)?$",
        flags=re.IGNORECASE,
    )
    for row in conn.execute(
        """
        SELECT nifti_file_series_id, file_name
        FROM nifti_file_series
        WHERE short_title = 'Vestibular-Schwannoma-MC-RC2'
        """
    ):
        match = pattern.match(row["file_name"] or "")
        if not match:
            continue
        patient_id, scan_date, scan_type = match.groups()
        update_nifti_series_fields(
            conn,
            row["nifti_file_series_id"],
            {
                "PatientID": patient_id,
                "StudyDate": scan_date,
                "SeriesDate": scan_date,
                "SeriesDescription": scan_type,
            },
            "dataset_specific:VS-MC-RC2_filename_convention",
        )
        updates += 1
    return updates


def apply_dataset_specific_nifti_enrichment(
    conn: sqlite3.Connection, files_dir: Path
) -> dict[str, int]:
    counts = {
        "rsna_brats2021_rows": apply_rsna_brats2021_mapping(conn, files_dir),
        "saros_rows": apply_saros_mapping(conn, files_dir),
        "breastdcedl_ispy2_rows": apply_breastdcedl_ispy2_mapping(conn, files_dir),
        "ct_org_rows": apply_ct_org_mapping(conn),
        "pancreas_ct_rows": apply_pancreas_ct_mapping(conn, files_dir),
        "brats_tcga_rows": apply_brats_tcga_mapping(conn),
        "brats_peds_rows": apply_brats_peds_mapping(conn),
        "plethora_rows": apply_plethora_mapping(conn),
        "vestibular_schwannoma_mc_rc2_rows": apply_vs_mc_rc2_mapping(conn),
    }
    conn.commit()
    write_meta(conn, "dataset_specific_enrichment_rows", counts)
    return counts


def apply_saros_external_source_references(conn: sqlite3.Connection) -> int:
    records = []
    for row in conn.execute(
        """
        SELECT
            d.derived_object_id,
            d.non_dicom_file_id,
            d.radiology_id,
            d.referenced_series_id,
            d.package_path,
            r.study_id,
            r.subject_id
        FROM derived_objects d
        JOIN radiology_series r
          ON r.radiology_id = d.radiology_id
        WHERE d.short_title = 'SAROS'
          AND NULLIF(d.referenced_series_id, '') IS NOT NULL
        """
    ):
        evidence = {
            "source": "Segmentation-Info_09-29-2023.csv",
            "tcia_case_id": row["subject_id"],
            "tcia_study_instance_uid": row["study_id"],
            "tcia_series_instance_uid": row["referenced_series_id"],
            "derived_package_path": row["package_path"],
        }
        records.append(
            (
                stable_hash_id("dor", row["derived_object_id"], row["referenced_series_id"], "saros_csv_source_series"),
                row["derived_object_id"],
                row["non_dicom_file_id"],
                row["radiology_id"],
                "",
                "",
                row["referenced_series_id"],
                "",
                "",
                "source_image_series",
                "saros_csv_source_series",
                "high",
                json_dumps(evidence),
            )
        )
    if records:
        conn.executemany(
            """
            INSERT OR REPLACE INTO derived_object_references
            (derived_object_reference_id, derived_object_id, derived_non_dicom_file_id,
             derived_radiology_id, referenced_non_dicom_file_id, referenced_radiology_id,
             referenced_series_id, referenced_file_name, referenced_package_path,
             reference_role, inference_method, confidence, evidence_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            records,
        )
        conn.commit()
    return len(records)


def apply_pancreas_ct_external_source_references(conn: sqlite3.Connection) -> int:
    records = []
    for row in conn.execute(
        """
        SELECT
            d.derived_object_id,
            d.non_dicom_file_id,
            d.radiology_id,
            d.file_name,
            d.referenced_series_id,
            d.package_path,
            r.study_id,
            r.subject_id
        FROM derived_objects d
        JOIN radiology_series r
          ON r.radiology_id = d.radiology_id
        WHERE d.short_title = 'Pancreas-CT'
          AND NULLIF(d.referenced_series_id, '') IS NOT NULL
        """
    ):
        evidence = {
            "source": f"{PANCREAS_CT_MANIFEST_FILE} plus IDC index",
            "tcia_patient_id": row["subject_id"],
            "tcia_study_instance_uid": row["study_id"],
            "tcia_series_instance_uid": row["referenced_series_id"],
            "derived_file_name": row["file_name"],
            "derived_package_path": row["package_path"],
        }
        records.append(
            (
                stable_hash_id(
                    "dor",
                    row["derived_object_id"],
                    row["referenced_series_id"],
                    "pancreas_ct_manifest_idc_patient_suffix",
                ),
                row["derived_object_id"],
                row["non_dicom_file_id"],
                row["radiology_id"],
                "",
                "",
                row["referenced_series_id"],
                "",
                "",
                "source_image_series",
                "pancreas_ct_manifest_idc_patient_suffix",
                "high",
                json_dumps(evidence),
            )
        )
    if records:
        conn.executemany(
            """
            INSERT OR REPLACE INTO derived_object_references
            (derived_object_reference_id, derived_object_id, derived_non_dicom_file_id,
             derived_radiology_id, referenced_non_dicom_file_id, referenced_radiology_id,
             referenced_series_id, referenced_file_name, referenced_package_path,
             reference_role, inference_method, confidence, evidence_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            records,
        )
        conn.commit()
    return len(records)


def write_canonical_layer_counts(conn: sqlite3.Connection) -> None:
    for table_name in (
        "nifti_file_series",
        "non_dicom_files",
        "radiology_series",
        "radiology_mr",
        "radiology_ct",
        "radiology_pet",
        "radiology_contrast",
        "derived_objects",
        "derived_object_references",
        "annotation_groups",
    ):
        write_meta(
            conn,
            f"{table_name}_rows",
            conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0],
        )


def postprocess_nifti_metadata(conn: sqlite3.Connection, files_dir: Path) -> dict[str, int]:
    counts: dict[str, int] = {}
    counts["direct_zip_nifti_package_files_inserted"] = inventory_cached_direct_zip_nifti_members(
        conn, files_dir
    )
    counts["local_extracted_package_files_inserted"] = inventory_local_extracted_package_files(
        conn, files_dir
    )
    build_nifti_file_series(conn)
    counts.update(apply_dataset_specific_nifti_enrichment(conn, files_dir))
    build_canonical_non_dicom_layer(conn)
    counts["saros_external_source_references"] = apply_saros_external_source_references(conn)
    counts["pancreas_ct_external_source_references"] = apply_pancreas_ct_external_source_references(
        conn
    )
    write_meta(conn, "postprocess_nifti_metadata_counts", counts)
    write_canonical_layer_counts(conn)
    return counts


def harvest(args: argparse.Namespace) -> None:
    source_conn = connect_source(Path(args.source_db))
    out_db = Path(args.out_db)
    files_dir = Path(args.files_dir)
    conn = open_output(out_db, replace=args.replace)

    nifti_rows = load_nifti_downloads(source_conn)
    insert_nifti_downloads(conn, nifti_rows)
    short_titles = sorted({row["short_title"] for row in nifti_rows if row["short_title"]})
    candidates = load_candidate_downloads(source_conn, short_titles)
    insert_candidate_downloads(conn, candidates)

    write_meta(
        conn,
        "generated_at_utc",
        dt.datetime.now(dt.timezone.utc).isoformat(),
    )
    write_meta(conn, "schema_version", SCHEMA_VERSION)
    write_meta(conn, "source_db", Path(args.source_db).name)
    write_meta(conn, "source_snapshot_meta", snapshot_meta(source_conn))
    write_meta(conn, "nifti_downloads", len(nifti_rows))
    write_meta(conn, "nifti_datasets", len(short_titles))
    write_meta(conn, "candidate_downloads", len(candidates))
    write_meta(conn, "args", public_harvest_args(args))

    if args.audit_only:
        return

    max_rows = None if args.max_rows_per_sheet < 0 else args.max_rows_per_sheet
    direct_candidates = [
        row
        for row in candidates
        if row["route"] in {"direct_spreadsheet", "direct_zip"}
        or extension_for_name(row.get("download_url") or "") in SPREADSHEET_EXTENSIONS | ZIP_EXTENSIONS
    ]
    if args.max_direct_downloads >= 0:
        direct_candidates = direct_candidates[: args.max_direct_downloads]
    for index, candidate in enumerate(direct_candidates, 1):
        print(
            f"[direct {index}/{len(direct_candidates)}] "
            f"{candidate['short_title']} {candidate['download_id']} {candidate['download_title']}",
            flush=True,
        )
        harvest_direct_candidate(
            conn,
            candidate,
            files_dir,
            args.max_zip_member_bytes,
            max_rows,
            args.reuse_cache,
            args.cache_only,
        )

    if args.include_aspera:
        aspera_candidates = [
            row
            for row in candidates
            if row["route"] == "aspera"
            and (
                not args.aspera_nifti_only
                or row["candidate_kind"] == "nifti_aspera_package"
            )
        ]
        if args.max_aspera_packages >= 0:
            aspera_candidates = aspera_candidates[: args.max_aspera_packages]
        for index, candidate in enumerate(aspera_candidates, 1):
            print(
                f"[aspera {index}/{len(aspera_candidates)}] "
                f"{candidate['short_title']} {candidate['download_id']} {candidate['download_title']}",
                flush=True,
            )
            harvest_aspera_candidate(
                conn,
                candidate,
                files_dir,
                args.ascli,
                args.max_aspera_metadata_files,
                max_rows,
                args.download_aspera_metadata_files,
                args.aspera_timeout_seconds,
                args.aspera_interactive_fallback,
                args.aspera_interactive_max_dirs,
                args.aspera_interactive_max_depth,
                args.aspera_interactive_descend_all,
                args.reuse_cache,
                args.cache_only,
            )

    import_sums_inventory(conn, Path(args.sums_inventory_tsv))
    import_quality_flags(conn, Path(args.quality_flags_tsv))
    postprocess_nifti_metadata(conn, files_dir)


def summary(db_path: Path) -> dict[str, Any]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    counts = {
        name: conn.execute(f"SELECT COUNT(*) AS c FROM {name}").fetchone()["c"]
        for name in [
            "nifti_downloads",
            "candidate_downloads",
            "package_files",
            "harvested_files",
            "tabular_sheets",
            "tabular_rows",
            "normalized_series_rows",
            "aspera_root_sums_inventory",
            "metadata_quality_flags",
            "nifti_file_series",
            "non_dicom_files",
            "radiology_series",
            "radiology_mr",
            "radiology_ct",
            "radiology_pet",
            "radiology_contrast",
            "derived_objects",
            "derived_object_references",
            "annotation_groups",
        ]
    }
    by_source = [
        dict(row)
        for row in conn.execute(
            """
            SELECT source_kind, status, COUNT(*) AS files
            FROM harvested_files
            GROUP BY source_kind, status
            ORDER BY source_kind, status
            """
        )
    ]
    normalized_fields = [
        dict(row)
        for row in conn.execute(
            """
            SELECT key AS field, COUNT(*) AS rows
            FROM normalized_series_rows, json_each(matched_columns_json)
            GROUP BY key
            ORDER BY rows DESC, field
            """
        )
    ]
    top_files = [
        dict(row)
        for row in conn.execute(
            """
            SELECT h.short_title, h.source_kind, h.file_name,
                   COUNT(n.normalized_row_id) AS normalized_rows
            FROM harvested_files h
            LEFT JOIN normalized_series_rows n ON n.file_id = h.file_id
            GROUP BY h.file_id
            ORDER BY normalized_rows DESC, h.short_title, h.file_name
            LIMIT 20
            """
        )
    ]
    return {
        "counts": counts,
        "harvested_files_by_source": by_source,
        "normalized_fields": normalized_fields,
        "top_normalized_files": top_files,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-db", default=str(DEFAULT_SOURCE_DB))
    parser.add_argument("--out-db", default=str(DEFAULT_OUT_DB))
    parser.add_argument("--files-dir", default=str(DEFAULT_OUT_DIR / "files"))
    parser.add_argument("--replace", action="store_true")
    parser.add_argument("--audit-only", action="store_true")
    parser.add_argument(
        "--reuse-cache",
        action="store_true",
        help="Reuse existing downloaded spreadsheets, ZIPs, and Aspera listing CSVs when present.",
    )
    parser.add_argument(
        "--cache-only",
        action="store_true",
        help="Do not download or browse; rebuild only from cached local artifacts.",
    )
    parser.add_argument("--include-aspera", action="store_true")
    parser.add_argument(
        "--aspera-nifti-only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Limit Aspera browsing to NIfTI-labeled package downloads by default.",
    )
    parser.add_argument(
        "--download-aspera-metadata-files",
        action="store_true",
        help="After browsing Aspera packages, download candidate spreadsheet members.",
    )
    parser.add_argument("--ascli", default=shutil.which("ascli") or "ascli")
    parser.add_argument(
        "--max-direct-downloads",
        type=int,
        default=-1,
        help="Limit direct URL downloads. -1 means no limit.",
    )
    parser.add_argument(
        "--max-aspera-packages",
        type=int,
        default=-1,
        help="Limit Aspera packages to browse. -1 means no limit.",
    )
    parser.add_argument(
        "--max-aspera-metadata-files",
        type=int,
        default=25,
        help="Maximum metadata files to download per Aspera package.",
    )
    parser.add_argument(
        "--aspera-timeout-seconds",
        type=int,
        default=300,
        help="Timeout for each ascli browse or receive command.",
    )
    parser.add_argument(
        "--aspera-interactive-fallback",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="On recursive browse failure, browse package paths incrementally.",
    )
    parser.add_argument(
        "--aspera-interactive-max-dirs",
        type=int,
        default=250,
        help="Maximum package directories to browse in interactive fallback mode.",
    )
    parser.add_argument(
        "--aspera-interactive-max-depth",
        type=int,
        default=2,
        help="Maximum directory depth for interactive fallback browsing.",
    )
    parser.add_argument(
        "--aspera-interactive-descend-all",
        action="store_true",
        help="In interactive fallback mode, descend all directories instead of metadata-named directories only.",
    )
    parser.add_argument(
        "--max-zip-member-bytes",
        type=int,
        default=50 * 1024 * 1024,
        help="Maximum metadata member size to extract from direct ZIP downloads.",
    )
    parser.add_argument(
        "--sums-inventory-tsv",
        default="",
        help="Optional TSV produced by check_aspera_root_sums.py to import package .sums inventory.",
    )
    parser.add_argument(
        "--quality-flags-tsv",
        default="",
        help="Optional TSV of confirmed source metadata quality flags.",
    )
    parser.add_argument(
        "--max-rows-per-sheet",
        type=int,
        default=-1,
        help="Maximum parsed rows per sheet. -1 means no limit.",
    )
    parser.add_argument("--summary", action="store_true")
    parser.add_argument(
        "--postprocess-only",
        action="store_true",
        help="Rebuild canonical NIfTI layers and apply cached dataset-specific enrichments without harvesting.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.summary:
        print(json.dumps(summary(Path(args.out_db)), indent=2, ensure_ascii=False))
        return 0
    if args.postprocess_only:
        conn = open_output(Path(args.out_db), replace=False)
        counts = postprocess_nifti_metadata(conn, Path(args.files_dir))
        print(json.dumps({"postprocess": counts, "summary": summary(Path(args.out_db))}, indent=2, ensure_ascii=False))
        return 0
    harvest(args)
    print(json.dumps(summary(Path(args.out_db)), indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
