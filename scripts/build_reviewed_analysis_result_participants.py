#!/usr/bin/env python3
"""Build reviewed participant memberships from official TCIA source files.

This developer utility intentionally does not download inputs. Pass a directory
containing the named TCIA downloads; the resulting compact CSV is consumed by
the V2 public non-DICOM builder and its JSON companion pins source digests.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path, PurePosixPath
from typing import Iterable

from openpyxl import load_workbook


REVIEWED_AT = "2026-08-27"
DATASET_TYPE = "Analysis Result"
SOURCES = {
    "DICOM-Glioma-SEG-metadata.zip": (
        "45839",
        "https://www.cancerimagingarchive.net/wp-content/uploads/TCGA_Segmentation-DCMQI-Metadata.zip",
    ),
    "ISBI-training.zip": (
        "45663",
        "https://www.cancerimagingarchive.net/wp-content/uploads/NCI-ISBI-2013-Prostate-Challenge-Training.zip",
    ),
    "ISBI-leaderboard.zip": (
        "45665",
        "https://www.cancerimagingarchive.net/wp-content/uploads/NCI-ISBI-2013-Prostate-Challenge-Leaderboard.zip",
    ),
    "ISBI-test.zip": (
        "45667",
        "https://www.cancerimagingarchive.net/wp-content/uploads/NCI-ISBI-2013-Prostate-Challenge-Test.zip",
    ),
    "LIDC-annotations.csv": (
        "55720",
        "https://www.cancerimagingarchive.net/wp-content/uploads/LIDC-annot-NLST501-annotations.csv",
    ),
    "LIDC-patients.csv": (
        "55734",
        "https://www.cancerimagingarchive.net/wp-content/uploads/LIDC-annot-NLST501-patients.csv",
    ),
    "MRQy-TCGA-GBM.xlsx": (
        "46111",
        "https://www.cancerimagingarchive.net/wp-content/uploads/TCGA_GBM_MRQy_Results.xlsx",
    ),
    "MRQy-TCGA-CESC.xlsx": (
        "46113",
        "https://www.cancerimagingarchive.net/wp-content/uploads/TCGA_CESCT1_and_T2_MRQy_Results.xlsx",
    ),
    "MRQy-CPTAC-GBM.xlsx": (
        "46115",
        "https://www.cancerimagingarchive.net/wp-content/uploads/CPTAC_GBM_MRQy_Results.xlsx",
    ),
    "TCGA-KIRC.csv": (
        "45577",
        "https://www.cancerimagingarchive.net/wp-content/uploads/FullDataforAnalysis071714.csv",
    ),
    "TCGA-OV-Proteogenomics.csv": (
        "45889",
        "https://www.cancerimagingarchive.net/wp-content/uploads/2020_02_26_HGSOC_proteomics.csv",
    ),
    "TCGA-OV-Radiogenomics-features.csv": (
        "45691",
        "https://www.cancerimagingarchive.net/wp-content/uploads/CompleteImageFeatureData012116.csv",
    ),
    "TCGA-OV-Radiogenomics-consensus.csv": (
        "45693",
        "https://www.cancerimagingarchive.net/wp-content/uploads/ImageFeatureConsensusMeasurements012116.csv",
    ),
    "TCGA-OV-Radiogenomics-clinical.csv": (
        "45695",
        "https://www.cancerimagingarchive.net/wp-content/uploads/ClinicalData101515.csv",
    ),
    "TCGA-OV-Radiogenomics-clovars.csv": (
        "45697",
        "https://www.cancerimagingarchive.net/wp-content/uploads/CLOVARScores121715.csv",
    ),
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_rows(
    rows: list[dict[str, str]], *, short_title: str, download_id: str,
    source_file: str, members: Iterable[tuple[str, str]], role: str,
) -> None:
    seen: set[str] = set()
    for raw_id, locator in members:
        participant_id = str(raw_id).strip()
        if not participant_id or participant_id in seen:
            continue
        seen.add(participant_id)
        rows.append({
            "dataset_type": DATASET_TYPE,
            "short_title": short_title,
            "download_id": download_id,
            "participant_id": participant_id,
            "raw_participant_id": participant_id,
            "participant_role": role,
            "source_file": source_file,
            "source_locator": locator,
        })


def csv_members(path: Path, column: str | None = None) -> list[tuple[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        field = column or str(reader.fieldnames[0])
        return [
            (str(row.get(field) or ""), f"row:{number}")
            for number, row in enumerate(reader, 2)
        ]


def workbook_members(path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    members: list[tuple[str, str]] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = [str(value or "").strip() for value in next(rows)]
        subject_column = next(
            index for index, value in enumerate(header) if value.casefold() == "subject id"
        )
        for row_number, row in enumerate(rows, 2):
            members.append((str(row[subject_column] or ""), f"sheet:{sheet.title};row:{row_number}"))
    return members


def build(source_dir: Path) -> tuple[list[dict[str, str]], dict[str, object]]:
    missing = [name for name in SOURCES if not (source_dir / name).is_file()]
    if missing:
        raise FileNotFoundError("Missing source files: " + ", ".join(missing))
    rows: list[dict[str, str]] = []

    name = "DICOM-Glioma-SEG-metadata.zip"
    with zipfile.ZipFile(source_dir / name) as archive:
        members = []
        for member in archive.namelist():
            match = re.search(r"TCGA-[A-Z0-9]{2}-[A-Z0-9]{4}", member)
            if match:
                members.append((match.group(0), f"zip:{member}"))
    add_rows(rows, short_title="DICOM-Glioma-SEG", download_id="45839", source_file=name,
             members=members, role="segmented_subject")

    for name, download_id in (
        ("ISBI-training.zip", "45663"),
        ("ISBI-leaderboard.zip", "45665"),
        ("ISBI-test.zip", "45667"),
    ):
        with zipfile.ZipFile(source_dir / name) as archive:
            members = []
            for member in archive.namelist():
                filename = PurePosixPath(member).name
                if filename.casefold().endswith(".nrrd"):
                    participant = re.sub(r"(?:_truth)?\.nrrd$", "", filename, flags=re.I)
                    members.append((participant, f"zip:{member}"))
        add_rows(rows, short_title="ISBI-MR-Prostate-2013", download_id=download_id,
                 source_file=name, members=members, role="imaged_subject")

    annotation_members = csv_members(source_dir / "LIDC-annotations.csv", "pid")
    for download_id, role in (("55710", "segmented_subject"), ("55720", "annotated_subject")):
        add_rows(rows, short_title="LIDC-annot-NLST501", download_id=download_id,
                 source_file="LIDC-annotations.csv", members=annotation_members, role=role)
    add_rows(rows, short_title="LIDC-annot-NLST501", download_id="55734",
             source_file="LIDC-patients.csv",
             members=csv_members(source_dir / "LIDC-patients.csv", "pid"), role="described_subject")

    for name, download_id in (
        ("MRQy-TCGA-GBM.xlsx", "46111"),
        ("MRQy-TCGA-CESC.xlsx", "46113"),
        ("MRQy-CPTAC-GBM.xlsx", "46115"),
    ):
        add_rows(rows, short_title="MRQy-Quality-Measures", download_id=download_id,
                 source_file=name, members=workbook_members(source_dir / name),
                 role="measured_subject")

    for name, short_title, download_id, column, role in (
        ("TCGA-KIRC.csv", "TCGA-KIRC-Radiogenomics", "45577", None, "annotated_subject"),
        ("TCGA-OV-Proteogenomics.csv", "TCGA-OV-Proteogenomics", "45889", "TCGA_ID", "described_subject"),
        ("TCGA-OV-Radiogenomics-features.csv", "TCGA-OV-Radiogenomics", "45691", "PatientID", "annotated_subject"),
        ("TCGA-OV-Radiogenomics-consensus.csv", "TCGA-OV-Radiogenomics", "45693", None, "annotated_subject"),
        ("TCGA-OV-Radiogenomics-clinical.csv", "TCGA-OV-Radiogenomics", "45695", None, "described_subject"),
        ("TCGA-OV-Radiogenomics-clovars.csv", "TCGA-OV-Radiogenomics", "45697", None, "described_subject"),
    ):
        add_rows(rows, short_title=short_title, download_id=download_id, source_file=name,
                 members=csv_members(source_dir / name, column), role=role)

    rows.sort(key=lambda row: (
        row["short_title"].casefold(), int(row["download_id"]), row["participant_id"]
    ))
    dataset_members: dict[str, set[str]] = defaultdict(set)
    download_members: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        dataset_members[row["short_title"]].add(row["participant_id"])
        download_members[f"{row['short_title']}:{row['download_id']}"].add(row["participant_id"])
    provenance = {
        "artifact": "reviewed_analysis_result_participants_v1",
        "reviewed_at": REVIEWED_AT,
        "review_status": "official_source_participant_inventory",
        "source_files": {
            name: {
                "download_id": download_id,
                "url": url,
                "sha256": sha256(source_dir / name),
                "size_bytes": (source_dir / name).stat().st_size,
            }
            for name, (download_id, url) in sorted(SOURCES.items())
        },
        "counts": {
            "membership_rows": len(rows),
            "participants_by_dataset": {
                key: len(value) for key, value in sorted(dataset_members.items())
            },
            "participants_by_download": {
                key: len(value) for key, value in sorted(download_members.items())
            },
        },
    }
    return rows, provenance


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--out-csv", type=Path, required=True)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()
    rows, provenance = build(args.source_dir)
    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.out_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(rows[0]),
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)
    provenance["inventory_sha256"] = sha256(args.out_csv)
    args.out_json.write_text(json.dumps(provenance, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(provenance["counts"], indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
